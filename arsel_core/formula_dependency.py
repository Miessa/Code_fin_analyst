# -*- coding: utf-8 -*-
"""Analyse bornée des dépendances de formules Excel pour le retrieval."""

import math
import re

from openpyxl.formula import Tokenizer
from openpyxl.utils.cell import range_boundaries, get_column_letter


FONCTIONS_AGREGATION = ("SUM", "SUMIF", "SUMIFS", "SUBTOTAL", "AGGREGATE")


def concept_attend_agregat(concept):
    structures = set(concept.get("structure_preferee") or [])
    return (
        concept.get("role_metrique") == "aggregate"
        or "scalar_aggregate" in structures
    )


def _separer_reference(reference, feuille_defaut):
    reference = reference.replace("$", "").strip()
    if "!" in reference:
        feuille, cellules = reference.rsplit("!", 1)
        feuille = feuille.strip("'").replace("''", "'")
    else:
        feuille, cellules = feuille_defaut, reference
    return feuille, cellules


class IndexDependancesFormules:
    def __init__(self, workbook, profondeur_max=3, references_max=250):
        self.workbook = workbook
        self.profondeur_max = profondeur_max
        self.references_max = references_max
        self._formules = {}
        self._analyses = {}

    def _formule(self, adresse):
        if adresse in self._formules:
            return self._formules[adresse]
        feuille, cellule = adresse.split("!", 1)
        try:
            valeur = self.workbook[feuille][cellule].value
        except Exception:
            valeur = None
        formule = valeur if isinstance(valeur, str) and valeur.startswith("=") else None
        self._formules[adresse] = formule
        return formule

    def _references(self, adresse, formule):
        feuille_courante = adresse.split("!", 1)[0]
        resultat = []
        try:
            tokens = Tokenizer(formule).items
        except Exception:
            return resultat
        for token in tokens:
            if token.type != "OPERAND" or token.subtype != "RANGE":
                continue
            feuille, cellules = _separer_reference(token.value, feuille_courante)
            if feuille not in self.workbook.sheetnames:
                continue
            try:
                min_col, min_row, max_col, max_row = range_boundaries(cellules)
            except (TypeError, ValueError):
                continue  # nom défini ou référence externe
            if None in (min_col, min_row, max_col, max_row):
                continue  # colonne/ligne entière ou nom ressemblant à une plage
            for ligne in range(min_row, max_row + 1):
                for colonne in range(min_col, max_col + 1):
                    resultat.append(f"{feuille}!{get_column_letter(colonne)}{ligne}")
                    if len(resultat) >= self.references_max:
                        return resultat
        return resultat

    def analyser(self, adresse):
        if not adresse or "!" not in adresse:
            return {"dependency_count": 0, "leaf_count": 0,
                    "has_aggregate_formula": False, "max_depth_reached": 0}
        if adresse in self._analyses:
            return dict(self._analyses[adresse])

        visites = set()
        feuilles = set()
        agregat = False
        profondeur_atteinte = 0

        def parcourir(courante, profondeur):
            nonlocal agregat, profondeur_atteinte
            if courante in visites or len(visites) >= self.references_max:
                return
            visites.add(courante)
            formule = self._formule(courante)
            if not formule:
                feuilles.add(courante)
                return
            profondeur_atteinte = max(profondeur_atteinte, profondeur)
            formule_maj = formule.upper()
            agregat = agregat or any(
                re.search(rf"\b{fonction}\s*\(", formule_maj)
                for fonction in FONCTIONS_AGREGATION
            )
            if profondeur >= self.profondeur_max:
                return
            for dependance in self._references(courante, formule):
                parcourir(dependance, profondeur + 1)

        parcourir(adresse, 0)
        resultat = {
            "dependency_count": max(0, len(visites) - 1),
            "leaf_count": len(feuilles),
            "has_aggregate_formula": agregat,
            "max_depth_reached": profondeur_atteinte,
        }
        self._analyses[adresse] = resultat
        return dict(resultat)


def calculer_boost_dependances(analyse, score_lexical, seuil_lexical=0.01):
    """Boost borné ; aucune pertinence lexicale implique aucun boost."""
    if score_lexical < seuil_lexical or analyse.get("dependency_count", 0) < 2:
        return 0.0
    boost = 0.025 * math.log2(1 + analyse["dependency_count"])
    if analyse.get("has_aggregate_formula"):
        boost += 0.04
    return min(0.20, boost)
