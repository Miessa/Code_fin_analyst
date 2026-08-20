# -*- coding: utf-8 -*-

"""
excel_tools.py

Outils Excel accessibles au LLM lorsqu'une décision
nécessite davantage de preuves.

Ces fonctions ne prennent aucune décision métier.
Elles retournent uniquement des faits provenant du classeur.
"""

from openpyxl.utils.cell import coordinate_from_string


def _separer_adresse(adresse: str):
    """
    Convertit 'InpC!F515' en ('InpC', 'F515').
    """

    if "!" not in adresse:
        raise ValueError(
            f"Adresse Excel invalide : {adresse}"
        )

    feuille, cellule = adresse.split("!", 1)

    return feuille, cellule


def lire_cellule(
    wb,
    adresse: str
) -> dict:
    """
    Lit la valeur calculée d'une cellule Excel.

    Args:
        wb: workbook openpyxl chargé avec data_only=True.
        adresse: Adresse Excel au format 'Feuille!Cellule'.

    Returns:
        Dictionnaire contenant l'adresse et la valeur.
    """

    feuille, cellule = _separer_adresse(
        adresse
    )

    if feuille not in wb.sheetnames:
        return {
            "adresse": adresse,
            "erreur": "feuille inexistante"
        }

    valeur = wb[feuille][cellule].value

    return {
        "adresse": adresse,
        "valeur": valeur
    }


def lire_formule(
    wb_formules,
    adresse: str
) -> dict:
    """
    Lit la formule Excel originale d'une cellule.

    Args:
        wb_formules: workbook openpyxl chargé avec data_only=False.
        adresse: Adresse Excel au format 'Feuille!Cellule'.

    Returns:
        Adresse et formule Excel éventuelle.
    """

    feuille, cellule = _separer_adresse(
        adresse
    )

    if feuille not in wb_formules.sheetnames:
        return {
            "adresse": adresse,
            "erreur": "feuille inexistante"
        }

    contenu = wb_formules[
        feuille
    ][cellule].value

    return {
        "adresse": adresse,
        "formule": (
            contenu
            if isinstance(contenu, str)
            and contenu.startswith("=")
            else None
        ),
        "contenu_brut": contenu
    }


def inspecter_voisinage(
    wb,
    adresse: str,
    rayon: int = 2
) -> list:
    """
    Retourne les cellules non vides autour d'une cellule.

    Args:
        wb: workbook openpyxl chargé avec data_only=True.
        adresse: Adresse Excel au format 'Feuille!Cellule'.
        rayon: Nombre de lignes/colonnes autour de la cellule.

    Returns:
        Liste des cellules non vides du voisinage.
    """

    feuille, cellule = _separer_adresse(
        adresse
    )

    if feuille not in wb.sheetnames:
        return []

    ws = wb[feuille]

    cell = ws[cellule]

    resultat = []

    min_row = max(
        1,
        cell.row - rayon
    )

    max_row = cell.row + rayon

    min_col = max(
        1,
        cell.column - rayon
    )

    max_col = cell.column + rayon

    for row in ws.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col
    ):

        for c in row:

            if c.value is not None:

                resultat.append({
                    "adresse":
                        f"{feuille}!{c.coordinate}",
                    "valeur": c.value
                })

    return resultat