# -*- coding: utf-8 -*-
"""
collecter_libelles.py — COLLECTE EXHAUSTIVE des libellés (rapide).

PERFORMANCE : en mode read_only, openpyxl est rapide pour lire les lignes EN
SÉQUENCE (iter_rows) mais très lent pour les accès dispersés (ws.cell). Ce
collecteur lit donc chaque ligne UNE fois, sous forme de tableau de valeurs,
et y trouve libellé + 1re valeur voisine sans aucun accès cellule ponctuel.
"""
import warnings; warnings.filterwarnings("ignore")
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FEUILLES_DEFAUT = [
    "InpC",
    "InpS",
    "InpS-M",
    "Summary",
    "Results",
    "Output",
    "Oper",
    "FS_Ann",
]


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _libelle_admis(texte, lmin=3, lmax=80):
    """Accepte notamment les acronymes financiers courts tels que EPC/IRR."""
    return lmin <= len(texte) < lmax


UNIT_RE = re.compile(
    r"(?:%|usd|eur|xaf|fcfa|cfa|gwh|mwh|kwh|mw|kw|years?|ans?|mois|months?|/kwh|/kw)",
    re.IGNORECASE,
)


def _contexte_superieur(lignes_precedentes, colonne, rayon=2):
    textes = []
    for ligne in reversed(lignes_precedentes[-8:]):
        for index in range(max(0, colonne - rayon), min(len(ligne), colonne + rayon + 1)):
            valeur = ligne[index]
            if isinstance(valeur, str) and valeur.strip() and valeur.strip() not in textes:
                textes.append(valeur.strip())
    unite = next((texte for texte in textes if UNIT_RE.search(texte)), None)
    return textes[:8], unite


def collecter(fichier, feuilles=FEUILLES_DEFAUT, max_lignes=None, lmin=3, lmax=80):
    """Catalogue : [{libelle, cellule_libelle, adresse_valeur, valeur}].
    Lecture 100% séquentielle : chaque ligne est convertie en liste de valeurs,
    puis on cherche dans cette liste (en mémoire) — pas d'accès cellule dispersé."""
    wv = load_workbook(fichier, data_only=True, read_only=True)
    catalogue = []
    for nom in feuilles:
        if nom not in wv.sheetnames:
            continue
        ws = wv[nom]
        dernier_texte_par_colonne = {}
        lignes_precedentes = []
        for ri, row in enumerate(ws.iter_rows(max_row=max_lignes), start=1):
            vals = [cell.value for cell in row]
            n = len(vals)
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    s = v.strip()
                    if _libelle_admis(s, lmin=lmin, lmax=lmax):
                        # 1re valeur numérique à droite, dans la MÊME ligne déjà lue
                        adr_val = val = None
                        for j in range(i + 1, n):
                            if _num(vals[j]):
                                adr_val = f"{nom}!{get_column_letter(j+1)}{ri}"
                                val = vals[j]
                                break
                        if adr_val is not None:
                            contexte_haut, unite_detectee = _contexte_superieur(
                                lignes_precedentes, j
                            )
                            voisins_textuels = [
                                str(vals[j]).strip()
                                for j in range(max(0, i - 4), min(n, i + 5))
                                if j != i
                                and isinstance(vals[j], str)
                                and vals[j].strip()
                            ]
                            catalogue.append({
                                "libelle": s,
                                "cellule_libelle": f"{nom}!{get_column_letter(i+1)}{ri}",
                                "adresse_valeur": adr_val,
                                "valeur": val,
                                "feuille": nom,
                                "section": dernier_texte_par_colonne.get(i),
                                "contexte": " | ".join(voisins_textuels),
                                "contexte_haut": " | ".join(contexte_haut),
                                "unite_detectee": unite_detectee,
                            })
            # Mettre à jour après la collecte afin que ``section`` désigne un
            # texte d'une ligne précédente, jamais le libellé lui-même.
            for i, valeur in enumerate(vals):
                if isinstance(valeur, str) and valeur.strip():
                    dernier_texte_par_colonne[i] = valeur.strip()
            lignes_precedentes.append(vals)
            if len(lignes_precedentes) > 8:
                lignes_precedentes.pop(0)
    return catalogue


if __name__ == "__main__":
    import sys, time
    f = sys.argv[1] if len(sys.argv) > 1 else "kikot.xlsm"
    t = time.time()
    cat = collecter(f)
    print(f"{len(cat)} libellés en {time.time()-t:.1f}s")


    for e in cat[:5]:
        print(f"  {e['cellule_libelle']} «{e['libelle'][:40]}» -> {e['adresse_valeur']}={e['valeur']}")
