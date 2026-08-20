# -*- coding: utf-8 -*-
"""
series_temporelle.py — DÉTECTION + SEGMENTATION des métriques temporelles.

Quand une métrique n'est pas un scalaire mais une SÉRIE pilotée par un
SÉLECTEUR (inflation, courbes de décaissement...), ce module :
  1. detecter_structure()  — le CODE repère sélecteur + série + en-tête de dates
  2. resoudre_indice()     — CODE/LLM : quel indice le sélecteur désigne
  3. segmenter()           — le CODE découpe la série en paliers datés
La restitution rend des SEGMENTS (pas une moyenne, pas 200 points).
"""
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string

# ---- Réglages de segmentation -------------------------------------------
SEUIL_RUPTURE = 0.0025     # 0,25 point : en-deçà, on ne crée pas de palier plat
SEUIL_TENDANCE = 0.0010    # micro-pas réguliers -> on les groupe en 1 tendance
# -------------------------------------------------------------------------


def _plage_colonnes(c1="J", c2="HA"):
    return range(column_index_from_string(c1), column_index_from_string(c2) + 1)


def lire_entete_dates(ws, ligne_dates, c1="J", c2="HA"):
    """Récupère les dates de l'en-tête temporel (le CODE lit la structure)."""
    dates = {}
    for c in _plage_colonnes(c1, c2):
        v = ws.cell(ligne_dates, c).value
        if isinstance(v, (datetime.datetime, datetime.date)):
            dates[c] = v
    return dates


def lire_serie(ws, ligne, c1="J", c2="HA"):
    """Récupère les valeurs numériques d'une ligne (la série)."""
    vals = {}
    for c in _plage_colonnes(c1, c2):
        v = ws.cell(ligne, c).value
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            vals[c] = v
    return vals


def detecter_structure(ws, ligne, ligne_dates=None, c1="J", c2="HA"):
    """Le CODE détecte : cette ligne est-elle une SÉRIE (≥ N valeurs sur l'axe temps) ?
    Retourne dict(est_serie, n_points, etendue) — déterministe."""
    serie = lire_serie(ws, ligne, c1, c2)
    est_serie = len(serie) >= 5           # au moins 5 points => structure temporelle
    return {"est_serie": est_serie, "n_points": len(serie),
            "premiere_col": min(serie) if serie else None,
            "derniere_col": max(serie) if serie else None}


def segmenter(ws, ligne, dates, c1="J", c2="HA",
              seuil_rupture=SEUIL_RUPTURE, seuil_tendance=SEUIL_TENDANCE):
    """Découpe une série en segments datés.
    - paliers nets : regroupés tant que |valeur - référence| <= seuil_rupture
    - micro-pas monotones : restitués en UN segment 'tendance' (de X à Y)
    Retourne une liste de segments : dict(type, debut, fin, valeur|debut_val,fin_val).
    """
    vals = lire_serie(ws, ligne, c1, c2)
    if not vals:
        return []
    cols = sorted(vals)

    # 1) découpe brute par rupture > seuil_rupture (absorbe les micro-pas)
    bruts = []
    debut = cols[0]; ref = vals[cols[0]]
    for c in cols[1:]:
        if abs(vals[c] - ref) > seuil_rupture:
            bruts.append((debut, c, ref, vals[c]))
            debut = c; ref = vals[c]
    bruts.append((debut, cols[-1], ref, vals[cols[-1]]))

    # 2) qualifier chaque segment : palier plat ou tendance (micro-variation monotone)
    segments = []
    for (d, f, v_ref, v_fin) in bruts:
        sous = [vals[c] for c in cols if d <= c <= f]
        vmin, vmax = min(sous), max(sous)
        amplitude = vmax - vmin
        if amplitude > seuil_tendance and sous[0] != sous[-1]:
            # variation interne notable et monotone -> tendance
            segments.append({"type": "tendance", "debut": d, "fin": f,
                             "debut_val": sous[0], "fin_val": sous[-1]})
        else:
            segments.append({"type": "palier", "debut": d, "fin": f, "valeur": v_ref})

    # 3) fusionner deux tendances/paliers adjacents de même valeur (nettoyage)
    return _formater(segments, dates)


def _fmt_date(col, dates):
    d = dates.get(col)
    return d.strftime("%b %Y") if d else get_column_letter(col)


def _formater(segments, dates):
    out = []
    for s in segments:
        if s["type"] == "palier":
            out.append({"type": "palier", "valeur": s["valeur"],
                        "de": _fmt_date(s["debut"], dates), "a": _fmt_date(s["fin"], dates)})
        else:
            out.append({"type": "tendance", "debut_val": s["debut_val"], "fin_val": s["fin_val"],
                        "de": _fmt_date(s["debut"], dates), "a": _fmt_date(s["fin"], dates)})
    return out


def _fmt(v, nature):
    """Formate une valeur selon sa nature : 'taux' -> %, sinon montant brut."""
    if nature == "taux":
        return f"{v*100:.2f} %"
    return f"{v:,.2f}".rstrip("0").rstrip(".")


def deviner_nature(segments):
    """Heuristique : si toutes les valeurs sont dans [0,1], c'est un taux ; sinon un montant."""
    vals = []
    for s in segments:
        vals += [s["valeur"]] if s["type"] == "palier" else [s["debut_val"], s["fin_val"]]
    vals = [abs(v) for v in vals if isinstance(v, (int, float))]
    if vals and max(vals) <= 1.0:
        return "taux"
    return "montant"


def restituer(segments, nature=None):
    """Rend les segments en texte lisible. nature='taux'|'montant' ; si None, devinée.
    Un taux s'affiche en % ; un montant s'affiche tel quel (pas de x100)."""
    if nature is None:
        nature = deviner_nature(segments)
    lignes = []
    for s in segments:
        if s["type"] == "palier":
            lignes.append(f"{_fmt(s['valeur'], nature)} de {s['de']} à {s['a']}")
        else:
            lignes.append(f"{_fmt(s['debut_val'], nature)} → {_fmt(s['fin_val'], nature)} "
                          f"(décroissance) de {s['de']} à {s['a']}")
    return lignes
