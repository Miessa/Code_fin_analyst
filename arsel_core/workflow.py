# -*- coding: utf-8 -*-
"""
arsel_analyse.py — SYSTÈME D'ANALYSE FINANCIÈRE ARSEL (point d'entrée unique)

Pipeline (option A : retrieval -> score -> rerank) :

  ÉTAPE 0  Concepts     — définitions + exemples
  ÉTAPE 1  Extraction   — pour chaque concept :
             (a) COLLECTE du catalogue (tous les libellés, une fois) ;
             (b) PRÉSÉLECTION déterministe (ontologie) -> ~6 candidats ;
             (c) DÉTECTEUR + FILTRE structurel -> score de compatibilité ;
             (d) GEMINI départage les candidats de tête (petit appel -> pas de 429) ;
             (e) contrôle de plausibilité.
  ÉTAPE 2  Validation   — l'analyste valide / corrige / saute
  ÉTAPE 3  Registre     — hypotheses_validees.json

Le code présélectionne et score (déterministe) ; le LLM ne fait qu'un rerank sur
quelques candidats. Sans clé : on prend le meilleur score.

Fichiers : arsel_analyse.py · referentiel_arsel.json (ontologie v0.5) ·
collecter_libelles.py · preselection.py · structure_detector.py ·
structure_filter.py · gemini_provider.py · resoudre.py · series_temporelle.py ·
primitives.py
Usage : python arsel_analyse.py MODELE.xlsm [ontologie.json]
"""
import sys, os, json, warnings,time
warnings.filterwarnings("ignore")


from openpyxl import load_workbook
from .collecter_libelles import collecter
from .structure_detector import caracteriser
from .structure_filter import diagnostiquer

from .resoudre import resoudre, resumer
from .semantic_selector import (
    OUTCOME_AMBIGUOUS,
    OUTCOME_NO_MATCH,
    OUTCOME_SELECTED,
    STATUS_NOT_REQUIRED,
    choisir_semantiquement,
)
from .gemini_provider import (
    disponible,
    appeler_json,
)
from .tfidf_search import IndexTfidf
from .llm_instrumentation import InstrumentationLLM
from .candidate_pipeline import recuperer_candidats, scorer_candidats
from .formula_dependency import IndexDependancesFormules

REFERENTIEL_DEFAUT = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "data",
    "referentiel_arsel.json",
)
CAT = {"factuel": "factuel (à reprendre)",
       "decisionnel": "décisionnel (levier — norme ARSEL)",
       "sortie": "métrique de sortie (à benchmarker)"}
ECART_SERRE = 0.15      # si top1 et top2 sont proches -> on demande à Gemini


def charger(chemin):
    if not os.path.exists(chemin):
        sys.exit(f"Référentiel introuvable : {chemin}")
    d = json.load(open(chemin, encoding="utf-8"))
    return d.get("entrees", []), d.get("_version", "?")


def plausible(v, lo, hi):
    return isinstance(v, (int, float)) and not isinstance(v, bool) and lo <= abs(v) <= hi


def doit_appeler_llm(scored):
    """
    Décide si Gemini doit intervenir.

    False :
        le meilleur candidat est suffisamment clair.

    True :
        ambiguïté, score faible ou signaux négatifs.
    """

    if not scored:
        return False

    # Un seul candidat : rien à départager
    if len(scored) == 1:
        return False

    top1 = scored[0]
    top2 = scored[1]

    score1 = top1.get("score", 0.0)
    score2 = top2.get("score", 0.0)

    ecart = score1 - score2

    negatifs = top1.get(
        "signaux_neg",
        []
    )

    # ------------------------------------------------------
    # Cas 1 : candidat très fort et clairement devant
    # ------------------------------------------------------

    if (
        score1 >= 0.90
        and ecart >= ECART_SERRE
        and not negatifs
    ):
        return False

    # ------------------------------------------------------
    # Cas 2 : les deux premiers sont proches
    # ------------------------------------------------------

    if ecart < ECART_SERRE:
        return True

    # ------------------------------------------------------
    # Cas 3 : même le meilleur candidat est assez faible
    # ------------------------------------------------------

    if score1 < 0.80:
        return True

    # ------------------------------------------------------
    # Cas 4 : le meilleur candidat a des contradictions
    # ------------------------------------------------------

    if negatifs:
        return True

    # ------------------------------------------------------
    # Sinon le déterministe suffit
    # ------------------------------------------------------

    return False

# ------------------------------------------------------------------- étape 0
def etape0(entrees):
    print("\n" + "═"*70 + "\nÉTAPE 0 — Concepts recherchés\n" + "═"*70)
    par_cat = {}
    for e in entrees:
        par_cat.setdefault(e.get("categorie", "autre"), []).append(e)
    for cat, lst in par_cat.items():
        print(f"\n── {CAT.get(cat, cat).upper()} ──")
        for e in lst:
            print(f"\n  • {e['cle']} — {e.get('description','')}")
            if e.get("definition"): print(f"      {e['definition']}")
    input("\n[Entrée] pour lancer l'extraction…")


# ------------------------------------------------------------------- étape 1
def etape1(fichier, entrees):
    print("\n" + "═"*70 + "\nÉTAPE 1 — Extraction (présélection · score · rerank)\n" + "═"*70)
    print("  Collecte des libellés…", end=" ", flush=True)

    t0 = time.time()

    catalogue = collecter(fichier)
    print(f"{len(catalogue)} libellés." f"({time.time() - t0:.1f}s)")
    index_tfidf = IndexTfidf(catalogue)

    # ==========================================================
    # DEBUG TEMPORAIRE — libellés liés à l'impôt
    # ==========================================================

    # print("\n===== LIBELLÉS TAX / CIT / INCOME =====")



    # for c in catalogue:

    #     lib = str(c.get("libelle", "")).lower()

    #     if any(
    #         terme in lib
    #         for terme in (
    #             "tax",
    #             "cit",
    #             "income",
    #             "impôt",
    #             "impot",
    #         )
    #     ):
    #         print(
    #             f"{c.get('cellule_libelle')} | "
    #             f"{c.get('libelle')} | "
    #             f"{c.get('adresse_valeur')} = "
    #             f"{c.get('valeur')}"
    #         )

    # print("======================================\n")

    # ==========================================================
    # DEBUG TEMPORAIRE — valeurs autour de 33 %
    # ==========================================================

    # print("\n===== DEBUG VALEURS AUTOUR DE 33% =====")

    # for c in catalogue:

    #     v = c.get("valeur")

    #     if (
    #         isinstance(v, (int, float))
    #         and not isinstance(v, bool)
    #         and 0.25 <= abs(v) <= 0.40
    #     ):
    #         print(
    #             f"{c.get('cellule_libelle')} | "
    #             f"{c.get('libelle')} | "
    #             f"{c.get('adresse_valeur')} = "
    #             f"{c.get('valeur')}"
    #         )

    # print("=======================================\n")
    llm_ok = disponible()
    instrumentation_llm = InstrumentationLLM(max_appels_par_metrique=1)
    #llm_ok = False
    print(
    f"  Gemini : "
    f"{'disponible' if llm_ok else 'indisponible'}\n"
)
    t0 = time.time()

    wb = load_workbook(
        fichier,
        data_only=True,
        # Le scoring fait des accès dispersés (cellules et lignes voisines).
        # En mode streaming, chaque accès peut rescanner une grande partie de
        # la feuille et faisait durer une seule métrique plus d'une minute.
        read_only=False
    )

    wb_formules = load_workbook(
    fichier,
    data_only=False,
    # Les dépendances de formules exigent des accès dispersés ; le mode
    # streaming rescannerait le classeur pour chaque cellule.
    read_only=False)
    index_dependances = IndexDependancesFormules(wb_formules)
    
    print(
        f"  Ouverture du modèle : "
        f"{time.time() - t0:.1f}s\n"
    )

    # wb = load_workbook(fichier, data_only=True, read_only=True)

    # ==========================================================
    # DEBUG TEMPORAIRE — Excel brut autour de 33 %
    # ==========================================================

    # print("\n===== DEBUG EXCEL BRUT AUTOUR DE 33% =====")

    # for ws in wb.worksheets:

    #     for row in ws.iter_rows():

    #         for cell in row:

    #             v = cell.value

    #             if (
    #                 isinstance(v, (int, float))
    #                 and not isinstance(v, bool)
    #                 and 0.25 <= abs(v) <= 0.40
    #             ):

    #                 voisins = []

    #                 # 3 cellules à gauche et à droite
    #                 debut = max(
    #                     0,
    #                     cell.column - 4
    #                 )

    #                 fin = min(
    #                     len(row),
    #                     cell.column + 3
    #                 )

    #                 for voisin in row[debut:fin]:

    #                     vv = voisin.value

    #                     if vv not in (
    #                         None,
    #                         ""
    #                     ):
    #                         voisins.append(
    #                             f"{voisin.coordinate}={vv}"
    #                         )

    #                 print(
    #                     f"{ws.title}!{cell.coordinate} = {v}"
    #                 )

    #                 print(
    #                     "    voisins : "
    #                     + " | ".join(voisins)
    #                 )

    # print("==========================================\n")


    resultats = []
    for e in entrees:
        t_metric = time.time()
        cle, desc, lo, hi = e["cle"], e["description"], *e["plage"]
        r = {"cle": cle, "categorie": e.get("categorie"), "description": desc,
             "definition": e.get("definition"), "nature": e.get("nature"),
             "unite": e.get("unite"), "resolver":e.get("resolver"),
             "adresse": None, "structure": None, "detail": None, "resume": None,
             "confiance": None, "signaux": None, "statut": None, "candidats": [],
             "selection_outcome": None, "execution_status": STATUS_NOT_REQUIRED,
             "llm": instrumentation_llm.pour_metrique(cle)}

        # (b) retrieval déterministe combiné : l'ancien lexical apporte de la
        # précision, TF-IDF apporte le rappel et les variantes terminologiques.
        cands, _, _ = recuperer_candidats(
            e, catalogue, index_tfidf,
            index_dependances=index_dependances,
        )
        if not cands:
            r["statut"] = "aucun candidat présélectionné"
            r["selection_outcome"] = OUTCOME_NO_MATCH
            resultats.append(r); print(f"  {cle:22s} → {r['statut']}"); continue

        # (c) détecteur + filtre : implémentation partagée avec le benchmark
        scored = scorer_candidats(wb, e, cands, wb_formules=wb_formules)
        # ==========================================================
        # DEBUG TEMPORAIRE — détail des candidats
        # ==========================================================

        # print(f"\n  ===== DEBUG {cle} =====")

        # for i, s in enumerate(scored[:5], start=1):

        #     diag = s.get("diag") or {}

        #     print(f"\n    Candidat {i}")
        #     print(
        #         f"      libellé        : "
        #         f"{s.get('libelle')}"
        #     )
        #     print(
        #         f"      cellule label  : "
        #         f"{s.get('cellule_libelle')}"
        #     )
        #     print(
        #         f"      cellule valeur : "
        #         f"{s.get('adresse_valeur')}"
        #     )
        #     print(
        #         f"      valeur         : "
        #         f"{s.get('valeur')}"
        #     )

        #     print(
        #         f"      score          : "
        #         f"{s.get('score'):.3f}"
        #     )

        #     print(
        #         f"      structure      : "
        #         f"{diag.get('structure')}"
        #     )

        #     print(
        #         f"      value_type     : "
        #         f"{diag.get('value_type')}"
        #     )

        #     print(
        #         f"      unit_family    : "
        #         f"{diag.get('unit_family')}"
        #     )

        #     print(
        #         f"      temporal_scope : "
        #         f"{diag.get('temporal_scope')}")
            
        #     print(
        #         f"      is_total       : "
        #         f"{diag.get('is_total')}"
        #     )

        #     print(
        #         f"      is_selector    : "
        #         f"{diag.get('is_selector')}"
        #     )

        #     print(
        #         f"      is_time_series : "
        #         f"{diag.get('is_time_series')}"
        #     )

        #     print(
        #         f"      signaux +      : "
        #         f"{s.get('signaux_pos')}"
        #     )

        #     print(
        #         f"      signaux -      : "
        #         f"{s.get('signaux_neg')}"
        #     )
        # r["candidats"] = [(s["cellule_libelle"], round(s["score"], 3)) for s in scored[:5]]

        # ==========================================================
        # (d) SÉLECTION SÉMANTIQUE
        # ==========================================================

        top = scored[0]
        decision_semantique = {
            "selection_outcome": OUTCOME_SELECTED,
            "execution_status": STATUS_NOT_REQUIRED,
            "cellule_libelle": top["cellule_libelle"],
            "confiance_semantique": None,
            "raison": "Sélection déterministe suffisamment claire.",
        }


        appel_llm = (
        llm_ok
        and doit_appeler_llm(scored)
        )

        print( f"    Routing LLM : "
           f"{'OUI' if appel_llm else 'NON'}")

        # print(
        #     f"\n    TOP STRUCTUREL   : "
        #     f"{top['cellule_libelle']} "
        #     f"score={top['score']:.3f}"
        # )

        # ----------------------------------------------------------
        # TEST AGENTIQUE TEMPORAIRE :
        # semantic selector uniquement pour is_taux
        # ----------------------------------------------------------

        if ( llm_ok and doit_appeler_llm(scored)):

            # print("    Semantic selector : OUI")

            decision_semantique = choisir_semantiquement(
                concept=e,
                candidats=scored[:8],
                appeler_llm_json=instrumentation_llm.instrumenter(
                    cle,
                    appeler_json,
                ),
            )

            # print(
            #     f"    Décision sémantique : "
            #     f"{decision_semantique.get('decision')}"
            # )

            # print(
            #     f"    Confiance sémantique: "
            #     f"{decision_semantique.get('confiance_semantique')}"
            # )

            # print(
            #     f"    Raison             : "
            #     f"{decision_semantique.get('raison')}"
            # )

            # ------------------------------------------------------
            # SELECT
            # ------------------------------------------------------

            if decision_semantique.get("selection_outcome") == OUTCOME_SELECTED:

                adr_choisie = decision_semantique.get(
                    "cellule_libelle"
                )

                candidat_choisi = next(
                    (
                        s for s in scored
                        if s["cellule_libelle"] == adr_choisie
                    ),
                    None
                )

                if candidat_choisi is not None:
                    top = candidat_choisi

            elif decision_semantique.get("selection_outcome") in {
                OUTCOME_AMBIGUOUS,
                OUTCOME_NO_MATCH,
            }:
                print(
                    f"    Sélection : "
                    f"{decision_semantique['selection_outcome']} / "
                    f"{decision_semantique['execution_status']} — "
                    "shortlist déterministe conservée pour validation"
                )

            # ------------------------------------------------------
            # EXPLORE
            # Aucun candidat suffisamment convaincant.
            # Pour ce premier test, on ne lance PAS encore
            # l'exploration agentique.
            # ------------------------------------------------------
            elif decision_semantique.get("decision") == "explore":

                # ======================================================
                # 1. Inspection approfondie des candidats existants
                #    avec Tool Calling Gemini + Excel
                # ======================================================

                # Exploration LLM historique supprimée : une ambiguïté
                # conserve désormais la shortlist déterministe pour l'analyste.
                decision_outils = None

                candidat_trouve_par_outils = False

                if decision_outils is not None and decision_outils.decision == "selected":

                    k = decision_outils.selected_candidate

                    if (
                        k is not None
                        and 1 <= k <= min(5, len(scored))
                    ):
                        top = scored[k - 1]

                        candidat_trouve_par_outils = True

                        print(
                            f"    Candidat retenu après inspection Excel : "
                            f"{top['cellule_libelle']}"
                        )

                        print(
                            f"    Confiance Gemini : "
                            f"{decision_outils.confidence:.0%}"
                        )

                        print(
                            f"    Raison : "
                            f"{decision_outils.reason}"
                        )

                # ======================================================
                # 2. Si Tool Calling n'a pas suffi :
                #    recherche de nouveaux candidats dans le catalogue
                # ======================================================

                if not candidat_trouve_par_outils:

                    nouveaux = []

                    # Aucun nouveau candidat
                    if not nouveaux:

                        r.update(
                            adresse=None,
                            structure=None,
                            detail={
                                "decision_semantique": decision_semantique
                            },
                            resume=(
                                "Aucun candidat supplémentaire trouvé "
                                "— exploration requise"
                            ),
                            confiance=decision_semantique.get(
                                "confiance_semantique"
                            ),
                            signaux={
                                "+": [],
                                "-": [
                                    decision_semantique.get(
                                        "raison",
                                        "exploration requise"
                                    )
                                ]
                            },
                            statut="à explorer"
                        )

                        print(
                            f"  {cle:22s} → "
                            f"à explorer                    "
                            f"{r['resume']}"
                        )

                        resultats.append(r)
                        continue

                # ======================================================
                # 2. Detector + filtre sur les nouveaux candidats
                # ======================================================

                    nouveaux_scored = []

                    for c in nouveaux:

                        try:

                            diag = caracteriser(
                                wb,
                                c["cellule_libelle"],
                                c["adresse_valeur"]
                            )

                            d = diagnostiquer(
                                diag,
                                e
                            )

                            score = d["score"]

                            neg = list(
                                d["signaux_negatifs"]
                            )

                            val = c.get("valeur")

                            if not plausible(
                                val,
                                lo,
                                hi
                            ):
                                score *= 0.3
                                neg.append(
                                    "valeur hors plage plausible"
                                )

                            nouveaux_scored.append({
                                **c,
                                "score": score,
                                "diag": diag,
                                "signaux_pos": d[
                                    "signaux_positifs"
                                ],
                                "signaux_neg": neg,
                            })

                        except Exception as ex:

                            print(
                                f"      ERREUR DIAGNOSTIC EXPLORE "
                                f"{c.get('cellule_libelle')} : "
                                f"{type(ex).__name__}: {ex}"
                            )

                            nouveaux_scored.append({
                                **c,
                                "score": 0.0,
                                "diag": None,
                                "signaux_pos": [],
                                "signaux_neg": [
                                    f"diagnostic échoué : "
                                    f"{type(ex).__name__}: {ex}"
                                ],
                            })


                    nouveaux_scored.sort(
                        key=lambda x: x["score"],
                        reverse=True
                    )


                # ======================================================
                # 3. DEBUG — candidats issus de l'exploration
                # ======================================================

                # print(
                #     "\n    ----- CANDIDATS APRÈS EXPLORATION -----"
                # )

                # for i, s in enumerate(
                #     nouveaux_scored[:10],
                #     start=1
                # ):

                #     diag = s.get("diag") or {}

                #     print(
                #         f"\n      Explore candidat {i}"
                #     )

                #     print(
                #         f"        libellé        : "
                #         f"{s.get('libelle')}"
                #     )

                #     print(
                #         f"        cellule label  : "
                #         f"{s.get('cellule_libelle')}"
                #     )

                #     print(
                #         f"        cellule valeur : "
                #         f"{s.get('adresse_valeur')}"
                #     )

                #     print(
                #         f"        valeur         : "
                #         f"{s.get('valeur')}"
                #     )

                #     print(
                #         f"        score          : "
                #         f"{s.get('score'):.3f}"
                #     )

                #     print(
                #         f"        structure      : "
                #         f"{diag.get('structure')}"
                #     )

                #     print(
                #         f"        unit_family    : "
                #         f"{diag.get('unit_family')}"
                #     )

                #     print(
                #         f"        signaux +      : "
                #         f"{s.get('signaux_pos')}"
                #     )

                #     print(
                #         f"        signaux -      : "
                #         f"{s.get('signaux_neg')}"
                #     )


                # ======================================================
                # 4. Deuxième et DERNIER semantic selector
                # ======================================================

                    decision_2 = {
                        "selection_outcome": OUTCOME_AMBIGUOUS,
                        "execution_status": STATUS_NOT_REQUIRED,
                        "raison": "Second sélecteur LLM supprimé.",
                    }

                # print(
                #     f"\n    Décision après exploration : "
                #     f"{decision_2.get('decision')}"
                # )

                # print(
                #     f"    Confiance après exploration: "
                #     f"{decision_2.get('confiance_semantique')}"
                # )

                # print(
                #     f"    Raison après exploration   : "
                #     f"{decision_2.get('raison')}"
                # )


                # ======================================================
                # 5A. Le second selector trouve enfin un candidat
                # ======================================================

                    if decision_2.get("decision") == "select":

                        adr_choisie = decision_2.get(
                            "cellule_libelle"
                        )

                        candidat_choisi = next(
                            (
                                s
                                for s in nouveaux_scored
                                if s.get(
                                    "cellule_libelle"
                                ) == adr_choisie
                            ),
                            None
                        )

                        if candidat_choisi is not None:

                            top = candidat_choisi

                            # print(
                            #     f"    Candidat retenu après exploration : "
                            #     f"{top['cellule_libelle']}"
                            # )

                        else:

                            # Cas défensif improbable
                            r.update(
                                adresse=None,
                                structure=None,
                                detail={
                                    "decision_semantique": decision_2
                                },
                                resume=(
                                    "Décision SELECT invalide "
                                    "après exploration"
                                ),
                                confiance=0.0,
                                signaux={
                                    "+": [],
                                    "-": [
                                        "candidat sélectionné introuvable"
                                    ]
                                },
                                statut="à explorer"
                            )

                            resultats.append(r)
                            continue


                # ======================================================
                # 5B. Toujours aucun bon candidat
                # ======================================================

                    else:

                        r.update(
                            adresse=None,
                            structure=None,
                            detail={
                                "decision_initiale": decision_semantique,
                                "decision_apres_exploration": decision_2,
                            },
                            resume=(
                                "Aucun candidat convaincant "
                                "même après exploration"
                            ),
                            confiance=decision_2.get(
                                "confiance_semantique"
                            ),
                            signaux={
                                "+": [],
                                "-": [
                                    decision_2.get(
                                        "raison",
                                        "exploration insuffisante"
                                    )
                                ]
                            },
                            statut="à explorer"
                        )

                        print(
                            f"  {cle:22s} → "
                            f"à explorer                    "
                            f"{r['resume']}"
                        )

                        resultats.append(r)
                        continue
        else:

            top = scored[0]

            # Pour les 3 autres métriques du test,
            # comportement déterministe actuel.
            # print("    Semantic selector : NON")


        # print(
        #     f"    TOP FINAL        : "
        #     f"{top['cellule_libelle']} "
        #     f"score={top['score']:.3f}"
        # )
        r["selection_outcome"] = decision_semantique.get("selection_outcome")
        r["execution_status"] = decision_semantique.get("execution_status")
        r["decision_semantique"] = decision_semantique
        r["llm"] = instrumentation_llm.pour_metrique(cle)

        # (e) résolution structure + plausibilité
        adr = top["cellule_libelle"]

        # print( f"    RESOLVER ONTOLOGIE : "
        # f"{e.get('resolver')}")
        try:
            res = resoudre(wb,adresse_libelle=adr,adresse_valeur=top.get("adresse_valeur"),nature=e.get("nature"),
            resolver=e.get("resolver"),catalogue=catalogue)

            # detail_debug = dict(
            #     res.get("detail", {})
            # )

            # if "segments" in detail_debug:
            #     detail_debug["segments"] = (
            #         f"{len(detail_debug['segments'])} segment(s)"
            #     )

            # if "restitution" in detail_debug:
            #     detail_debug["restitution"] = (
            #         detail_debug["restitution"][:3]
            #     )

            # print(
            #     f"    RESOLUTION      : "
            #     f"{detail_debug.get('resolution')}"
            # )

            # print(
            #     f"    STRUCTURE FINALE: "
            #     f"{res.get('structure')}"
            # )

            # print(
            #     f"    DETAIL FINAL    : "
            #     f"{detail_debug}"
            # )
            statut = "proposé"
            dv = res["detail"].get("valeur")
            if dv is not None and not plausible(dv, lo, hi):
                statut = "proposé (hors plage — douteux)"
            r.update(adresse=adr, structure=res["structure"], detail=res["detail"],
                     resume=resumer(res), confiance=top["score"],
                     signaux={"+": top["signaux_pos"], "-": top["signaux_neg"]}, statut=statut)
        except Exception as ex:

            print(
                f"    ERREUR RESOLUTION "
                f"{cle} / {adr} : "
                f"{type(ex).__name__}: {ex}"
            )

            r.update(
                adresse=adr,
                detail={
                    "valeur": top.get("valeur"),
                    "erreur_resolution": (
                        f"{type(ex).__name__}: {ex}"
                    )
                },
                resume=(
                    f"{top.get('adresse_valeur')} "
                    f"= {top.get('valeur')}"
                ),
                confiance=top["score"],
                statut="proposé (structure simple)"
            )

        conf = f"[{r['confiance']:.0%}]" if r["confiance"] is not None else ""
        print(f"  {cle:22s} → "f"{r['statut']:30s} "f"{conf:6s} "f"{r['resume'] or ''} "f"({time.time() - t_metric:.1f}s)")
        resultats.append(r)
    resume_llm = instrumentation_llm.resume()
    print(
        "\n  Appels LLM : "
        f"{resume_llm['total_calls']} | "
        f"succès={resume_llm['total_successes']} | "
        f"échecs={resume_llm['total_failures']} | "
        f"latence={resume_llm['latency_ms']:.0f} ms"
    )
    return (
    resultats,
    wb,
    wb_formules,
    catalogue
)


# ------------------------------------------------------------------- étape 2
def etape2(wb, resultats, catalogue):
    print(
        "\n" + "═"*70 +
        "\nÉTAPE 2 — Validation par l'analyste\n" +
        "═"*70
    )

    print(
        "  [v] valider · "
        "[a] ajouter/modifier valeur · "
        "[c] corriger l'adresse · "
        "[s] indisponible · "
        "[q] quitter\n"
    )

    valides = []

    for r in resultats:

        print("─"*60)

        print(
            f"  {r['cle']}  "
            f"({CAT.get(r['categorie'], r['categorie'])})"
        )

        print(
            f"  {r['description']}"
        )

        # ======================================================
        # Affichage de la proposition automatique
        # ======================================================

        if r.get("adresse"):

            confiance = (
                f"{r['confiance']:.0%}"
                if r.get("confiance") is not None
                else "—"
            )

            print(
                f"  STRUCTURE : {r.get('structure')}   "
                f"CONFIANCE : {confiance}"
            )

            print(
                f"  PROPOSITION : "
                f"{r.get('resume') or '—'}"
            )

            if r.get("signaux"):

                if r["signaux"].get("+"):
                    print(
                        f"    + "
                        f"{', '.join(r['signaux']['+'])}"
                    )

                if r["signaux"].get("-"):
                    print(
                        f"    - "
                        f"{', '.join(r['signaux']['-'])}"
                    )

            if (
                r.get("candidats")
                and len(r["candidats"]) > 1
            ):

                autres = ", ".join(
                    f"{a}({s})"
                    for a, s in r["candidats"][1:]
                )

                print(
                    f"    autres candidats : "
                    f"{autres}"
                )

        else:

            print(
                f"  PROPOSITION : "
                f"{r.get('resume') or r.get('statut') or 'aucune'}"
            )


        # ======================================================
        # Choix analyste
        # ======================================================

        rep = input(
            "  [v/a/c/s/q] > "
        ).strip().lower()


        # ======================================================
        # QUITTER
        # ======================================================

        if rep == "q":
            break


        # ======================================================
        # SAISIE / MODIFICATION MANUELLE
        # ======================================================

        if rep == "a":

            valeur_txt = input(
                "  valeur correcte > "
            ).strip()

            # Autoriser virgule décimale française
            valeur_normalisee = valeur_txt.replace(",", ".")

            try:
                valeur = float(
                    valeur_normalisee
                )

            except ValueError:
                valeur = valeur_txt

            r = {
                **r,
                "adresse": None,
                "structure": "manual",
                "detail": {
                    "valeur": valeur,
                    "source": "analyste",
                },
                "resume": (
                    f"{valeur} "
                    f"(saisie analyste)"
                ),
                "confiance": 1.0,
                "signaux": {
                    "+": [
                        "valeur fournie par l'analyste"
                    ],
                    "-": [],
                },
                "statut": (
                    "saisi manuellement (analyste)"
                ),
            }

            print(
                f"  → valeur retenue : "
                f"{valeur}"
            )

            valides.append(r)
            continue


        # ======================================================
        # CORRECTION DE L'ADRESSE EXCEL
        # ======================================================

        if rep == "c":

            adr = input(
                "  adresse LIBELLÉ correcte "
                "(Feuille!Cellule) > "
            ).strip()

            try:

                res = resoudre(
                    wb,
                    adresse_libelle=adr,
                    nature=r.get("nature"),
                    resolver=r.get("resolver"),
                    catalogue=catalogue,
                )

                r = {
                    **r,
                    "adresse": adr,
                    "structure": res["structure"],
                    "detail": res["detail"],
                    "resume": resumer(res),
                    "statut": (
                        "corrigé (analyste)"
                    ),
                }

            except Exception as ex:

                r = {
                    **r,
                    "adresse": adr,
                    "resume": (
                        f"(non résolu : {ex})"
                    ),
                    "statut": (
                        "corrigé (analyste)"
                    ),
                }

            print(
                f"  → {r['resume']}"
            )

            valides.append(r)
            continue


        # ======================================================
        # MARQUER COMME INDISPONIBLE
        # ======================================================

        if rep == "s":

            r = {
                **r,
                "adresse": None,
                "structure": None,
                "detail": {
                    "valeur": None,
                    "source": "indisponible",
                },
                "resume": "non disponible",
                "confiance": None,
                "statut": "non disponible",
            }

            print(
                "  → métrique conservée "
                "comme non disponible"
            )

            valides.append(r)
            continue


        # ======================================================
        # VALIDATION
        # ======================================================

        if rep == "v":

            r = {
                **r,
                "statut": "validé (analyste)"
            }

            valides.append(r)
            continue


        # ======================================================
        # Réponse invalide
        # ======================================================
# ======================================================
# Si l'utilisateur tape directement une valeur
# on l'interprète comme une saisie manuelle
# ======================================================

        valeur_txt = rep

        try:
            valeur = float(
                valeur_txt
                .replace(" ", "")
                .replace(",", ".")
            )

            r = {
                **r,
                "adresse": None,
                "structure": "manual",
                "detail": {
                    "valeur": valeur,
                    "source": "analyste",
                },
                "resume": (
                    f"{valeur} "
                    f"(saisie analyste)"
                ),
                "confiance": 1.0,
                "signaux": {
                    "+": [
                        "valeur fournie par l'analyste"
                    ],
                    "-": [],
                },
                "statut": "saisi manuellement (analyste)",
            }

            print(
                f"  → valeur retenue : {valeur}"
            )

            valides.append(r)

        except ValueError:

            print(
                "  → choix invalide. "
                "Utilisez v, a, c, s ou q."
            )

            # IMPORTANT :
            # ne pas transformer automatiquement
            # la métrique en indisponible
            continue

    return valides


# ------------------------------------------------------------------- étape 3
def etape3(valides, sortie="hypotheses_validees.json"):
    print(
        "\n" + "═"*70 +
        "\nÉTAPE 3 — Registre des hypothèses validées\n" +
        "═"*70
    )

    registre = []

    for r in valides:
        detail = r.get("detail") or {}

        # ==========================================================
        # Récupération uniforme de la valeur métier
        # ==========================================================

        if detail.get("valeur") is not None:
            # Cas scalaire
            valeur = detail.get("valeur")

        elif detail.get("restitution"):
            # Cas série temporelle déjà résumée par le resolver
            valeur = detail.get("restitution")

        elif detail.get("segments"):
            # Repli : conserver directement les segments
            valeur = detail.get("segments")

        else:
            valeur = None


        # ==========================================================
        # 2. Déterminer le type de valeur
        # ==========================================================

        if detail.get("valeur") is not None:
            type_valeur = "scalar"

        elif detail.get("segments") or detail.get("restitution"):
            type_valeur = "time_series"

        else:
            type_valeur = "missing"


        # ------------------------------------------------------
        # Déterminer la source de la valeur
        # ------------------------------------------------------

        source = detail.get("source")

        if not source:

            statut = r.get("statut", "")

            if "manuellement" in statut:
                source = "analyste"

            elif "corrigé" in statut:
                source = "analyste"

            elif "validé" in statut:
                source = "extraction"

            elif statut == "non disponible":
                source = "indisponible"

            else:
                source = "extraction"


        # ------------------------------------------------------
        # Construire l'entrée métier propre
        # ------------------------------------------------------

        entree = {
        "cle": r.get("cle"),
        "description": r.get("description"),
        "categorie": r.get("categorie"),

        "valeur": valeur,
        "type_valeur": type_valeur,

        "nature": r.get("nature"),
        "unite": r.get("unite"),
        "adresse": r.get("adresse"),
        "source": source,
        "confiance": r.get("confiance"),
        "statut": r.get("statut"),
        "selection_outcome": r.get("selection_outcome"),
        "execution_status": r.get("execution_status"),
        "llm": r.get("llm"),
    }

        registre.append(entree)


    # ----------------------------------------------------------
    # Écriture JSON
    # ----------------------------------------------------------

    with open(
        sortie,
        "w",
        encoding="utf-8"
    ) as f:

        json.dump(
            registre,
            f,
            ensure_ascii=False,
            indent=2,
            default=str
        )


    # ----------------------------------------------------------
    # Résumé console
    # ----------------------------------------------------------

    print(
        f"  {len(registre)} métrique(s) "
        f"→ {sortie}\n"
    )

    for r in registre:

        valeur = (
            r["valeur"]
            if r["valeur"] is not None
            else "NON DISPONIBLE"
        )

        source = r.get(
            "source",
            "—"
        )

        print(
            f"  {r['cle']:22s} "
            f"{str(valeur):18s} "
            f"[{source}]"
        )

    return sortie

def main():
    fichier = sys.argv[1] if len(sys.argv) > 1 else "kikot.xlsm"
    chemin = sys.argv[2] if len(sys.argv) > 2 else REFERENTIEL_DEFAUT
    if not os.path.exists(fichier):
        sys.exit(f"Modèle introuvable : {fichier}")
    entrees, version = charger(chemin)

     # ==========================================================
    # DEBUG TEMPORAIRE — seulement 4 métriques
    # ==========================================================
    # CLES_DEBUG = {
    #     "is_taux",
    #     "dscr_cible",
    #     "productible",
    #     "cout_construction",
    # }

    # entrees = [
    #     e for e in entrees
    #     if e["cle"] in CLES_DEBUG
    # ]

    
    
    print(f"\nOntologie v{version} — {len(entrees)} concepts | Modèle : {fichier}")
    etape0(entrees)
    (resultats,wb,wb_formules,catalogue) = etape1( fichier, entrees)
    valides = etape2(wb, resultats, catalogue)
    registre_path = etape3(valides)

    # Phase 2 consumes only the analyst-validated registry. It never reads the
    # workbook and therefore cannot alter Phase 1 extraction results.
    from phase2.pipeline import executer_phase2
    contexte_phase2 = None
    chemin_contexte = next(
        (p for p in ("phase2_context.json", os.path.join("phase2", "phase2_context.json"))
         if os.path.exists(p) and os.path.getsize(p) > 0),
        None,
    )
    if chemin_contexte:
        with open(chemin_contexte, encoding="utf-8") as f:
            contexte_phase2 = json.load(f)
    _, sortie_json, sortie_md = executer_phase2(
        registre_path,
        contexte=contexte_phase2,
    )
    print("\nPHASE 2 — Analyse financière et benchmarks")
    print(f"  JSON     → {sortie_json}")
    print(f"  Markdown → {sortie_md}")


if __name__ == "__main__":
    main()
