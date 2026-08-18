# -*- coding: utf-8 -*-
"""
resoudre.py — RÉSOLUTION D'UNE MÉTRIQUE selon sa structure détectée.

Prend une cellule (pointée par le LLM ou l'analyste), détecte sa structure via
les primitives, et applique le traitement adapté :
  • valeur        -> lit le scalaire
  • serie         -> segmente en paliers datés
  • selection -> lit le sélecteur ; certains resolvers spécialisés
               peuvent suivre l'option active jusqu'à sa donnée cible
  • decomposition -> renvoie le total + les composantes
La sortie est uniforme : dict(structure, valeur|segments|composantes, provenance).
Le CODE lit et structure ; aucune valeur n'est inventée.
"""
from openpyxl.utils import column_index_from_string, get_column_letter
from primitives import (
    caracteriser,
    detecter_serie,
    detecter_selection,
    trouver_entete_dates,
)

from series_temporelle import (
    lire_entete_dates,
    segmenter,
    restituer,
    deviner_nature,
)


def _coord(adresse):
    feuille, cell = adresse.split("!", 1)
    col = "".join(ch for ch in cell if ch.isalpha())
    row = int("".join(ch for ch in cell if ch.isdigit()))
    return feuille, column_index_from_string(col), row

def trouver_options_selecteur(
    ws,
    r,
    c,
    indice,
    rayon_lignes=40,
    rayon_colonnes=20
):
    """
    Cherche autour du sélecteur des couples :
        entier -> texte

    Retourne des candidats enrichis :
        {
            "label": "Cameroon CPI",
            "address": "InpC!G123",
            "distance": 4
        }

    Les résultats sont triés du plus proche au plus éloigné.
    """

    resultats = []

    r1 = max(1, r - rayon_lignes)
    r2 = min(ws.max_row, r + rayon_lignes)

    c1 = max(1, c - rayon_colonnes)
    c2 = min(ws.max_column, c + rayon_colonnes)

    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):

            v = ws.cell(rr, cc).value

            if not isinstance(v, (int, float)):
                continue

            if not float(v).is_integer():
                continue

            if int(v) != int(indice):
                continue

            distance = abs(rr - r) + abs(cc - c)

            # texte immédiatement à droite
            if cc < ws.max_column:
                texte = ws.cell(rr, cc + 1).value

                if isinstance(texte, str) and texte.strip():
                    resultats.append({
                        "label": texte.strip(),
                        "address": f"{ws.title}!{get_column_letter(cc + 1)}{rr}",
                        "distance": distance
                    })

            # texte immédiatement à gauche
            if cc > 1:
                texte = ws.cell(rr, cc - 1).value

                if isinstance(texte, str) and texte.strip():
                    resultats.append({
                        "label": texte.strip(),
                        "address": f"{ws.title}!{get_column_letter(cc - 1)}{rr}",
                        "distance": distance
                    })

    # dédoublonnage par label + adresse
    uniques = {}
    for item in resultats:
        key = (item["label"].lower(), item["address"])
        uniques[key] = item

    resultats = list(uniques.values())

    # priorité aux options les plus proches du sélecteur
    resultats.sort(key=lambda x: x["distance"])

    return resultats

def chercher_libelle_catalogue(
    catalogue,
    texte
):
    if not catalogue or not texte:
        return []

    cible = texte.lower().strip()

    exacts = []
    partiels = []

    for candidat in catalogue:

        libelle = (
            candidat.get("libelle")
            or ""
        ).lower().strip()

        if not libelle:
            continue

        if libelle == cible:
            exacts.append(candidat)

        elif cible in libelle:
            partiels.append(candidat)

    # priorité aux correspondances exactes
    return exacts + partiels

def _resoudre_generique(wb, adresse_libelle, nature=None):
    """adresse_libelle = 'Feuille!Cellule' du LIBELLÉ de la métrique.
    nature = 'taux' | 'montant' (attendu par le concept ; sinon deviné).
    Cascade de priorité :
      1) TOTAL déjà calculé (décomposition avec total)   -> on le prend
      2) SÉRIE (pas de total)   -> segmentation, avec la bonne nature
      3) SÉLECTION              -> option active (indice résolu ensuite)
      4) VALEUR simple
    """
    feuille, c, r = _coord(adresse_libelle)
    ws = wb[feuille]
    prof = caracteriser(ws, r, c)
    types = prof["types"]
    resultat = {"cellule_libelle": adresse_libelle,
                "libelle": prof.get("libelle"),
                "structure": types, "detail": {}, "nature": nature}

    # 1) PRIORITÉ AU TOTAL : si décomposition avec total, on le retient d'abord
    total = None
    if "decomposition" in types:
        dec = next(p for p in prof["primitives"] if p["type"] == "decomposition")
        resultat["detail"]["composantes"] = dec["composantes"]
        total = dec.get("total")
        resultat["detail"]["total"] = total

    # 2) SÉRIE — seulement si PAS de total exploitable
    if "serie" in types and not total:
        ligne_dates = next((p.get("ligne_dates") for p in prof["primitives"]
                            if p["type"] == "serie"), None)
        dates = lire_entete_dates(ws, ligne_dates) if ligne_dates else {}
        segs = segmenter(ws, r, dates)
        nat = nature or deviner_nature(segs)   # taux -> %, montant -> brut
        resultat["detail"]["segments"] = segs
        resultat["detail"]["restitution"] = restituer(segs, nat)
        resultat["nature"] = nat

    # 3) SÉLECTION
    if "selection" in types:
        sel = next(p for p in prof["primitives"] if p["type"] == "selection")
        resultat["detail"]["selecteur"] = sel["selecteur"]
        resultat["detail"]["option_active"] = sel["valeur_active"]

    # 4) VALEUR simple
    if "valeur" in types:
        val = next(p for p in prof["primitives"] if p["type"] == "valeur")
        resultat["detail"]["valeur"] = val["valeur"]
        resultat["detail"]["adresse_valeur"] = val["adresse"]

    return resultat

def resoudre_valeur_directe(
    wb,
    adresse_libelle,
    adresse_valeur,
    nature=None
):
    """
    Résolution directe d'une métrique scalaire déjà identifiée
    par le pipeline.

    IMPORTANT :
    ne rescane pas le voisinage de la cellule ;
    utilise exactement adresse_valeur fournie par le candidat.
    """

    feuille_label, c_label, r_label = _coord(adresse_libelle)
    ws_label = wb[feuille_label]

    libelle = ws_label.cell(r_label, c_label).value

    if not adresse_valeur:
        return {
            "cellule_libelle": adresse_libelle,
            "libelle": libelle,
            "structure": [],
            "nature": nature,
            "detail": {
                "resolution": "direct_value_failed",
                "reason": "adresse_valeur absente"
            }
        }

    feuille_val, c_val, r_val = _coord(adresse_valeur)
    ws_val = wb[feuille_val]

    valeur = ws_val.cell(r_val, c_val).value

    return {
        "cellule_libelle": adresse_libelle,
        "cellule_source": adresse_valeur,
        "libelle": libelle,
        "structure": ["valeur"],
        "nature": nature,
        "detail": {
            "resolution": "direct_value",
            "valeur": valeur,
            "adresse_valeur": adresse_valeur
        }
    }


def resoudre_index_selector_to_series(
    wb,
    adresse_libelle,
    adresse_valeur=None,
    nature="taux",
    catalogue=None
):
    """
    Résout une métrique indexée/inflation.

    Cas A :
        le candidat est directement une série temporelle
        (ex. "Cameroon CPI %")
        -> retourner la série.

    Cas B :
        le candidat est un sélecteur
        (ex. valeur 2 = Cameroon CPI)
        -> résoudre l'option
        -> retrouver la ligne correspondante
        -> retourner sa série.

    Aucun mapping spécifique à Kikot n'est codé ici.
    """
    
    feuille, c, r = _coord(adresse_libelle)
    ws = wb[feuille]
    nature = nature or "taux"
    libelle = ws.cell(r, c).value

    # ==========================================================
    # CAS A — LE CANDIDAT EST DÉJÀ UNE SÉRIE
    # ==========================================================

    serie = detecter_serie(ws, r)

    if serie.get("present"):

        ligne_dates = trouver_entete_dates(ws, r)

         # Trace de diagnostique:
          # une serie a été détectée,mais aucun en-tête temporel
           # n'a pu être identifié

        warning = None

        if not ligne_dates:
            warning = "serie_detectee_sans_entete_temporel"
        

        dates = {}
        if ligne_dates:
            dates = lire_entete_dates(ws, ligne_dates)

        segments = segmenter(
            ws,
            r,
            dates
        )

        return {
            "cellule_libelle": adresse_libelle,
            "cellule_source": adresse_libelle,
            "libelle": libelle,

            "structure": ["serie"],
            "nature": nature,

            "detail": {
                "resolution": "direct_series",
                "warning": warning,
                "ligne_dates": ligne_dates,
                "segments": segments,
                "restitution": restituer(
                    segments,
                    nature
                )
            }
        }

    # ==========================================================
    # CAS B — LE CANDIDAT EST UN SÉLECTEUR
    # ==========================================================

    indice =None
    selection =None

    if adresse_valeur:
        try:
            f_val, c_val, r_val = _coord(adresse_valeur)
            v = wb[f_val].cell(r_val, c_val).value
            if isinstance(v, (int, float)) and not isinstance(v,bool):
                if float(v).is_integer():
                    indice = int(v)
        except Exception:
            pass
    if indice is None:
        selection = detecter_selection(ws,r,c)
        if selection.get("present"):
            indice = selection.get("valeur_active")

    

    if indice is None:
        return {
            "cellule_libelle": adresse_libelle,
            "libelle": libelle,
            "structure": [],
            "nature": nature,
            "detail": {
                "resolution": "unresolved_selector",
                "reason": "Le candidat n'est ni une série temporelle, ni un selecteur résoluble."
            }
        }

    # ----------------------------------------------------------
    # B1. Résoudre indice -> libellé
    # ----------------------------------------------------------

    options = trouver_options_selecteur(
        ws,
        r,
        c,
        indice
    )

    if not options:
        return {
            "cellule_libelle": adresse_libelle,
            "libelle": libelle,
            "structure": ["selection"],
            "nature": nature,
            "detail": {
                "resolution": "unresolved_selector",
                "selector_value": indice,
                "reason": (
                    "Aucune option correspondant au sélecteur "
                    "n'a été trouvée."
                )
            }
        }

    # Plusieurs textes peuvent éventuellement être trouvés.
    # On tente chacun jusqu'à trouver une vraie série.
    for option_info in options:

        option = option_info["label"]

        # ------------------------------------------------------
        # B2. Retrouver cette option dans le catalogue
        # ------------------------------------------------------

        candidats = chercher_libelle_catalogue(
            catalogue,
            option
        )

        for candidat in candidats:

            adresse_cible = candidat.get(
                "cellule_libelle"
            )

            if not adresse_cible:
                continue

            f2, c2, r2 = _coord(adresse_cible)
            ws2 = wb[f2]

            # --------------------------------------------------
            # B3. Vérifier que la cible est réellement une série
            # --------------------------------------------------

            serie_cible = detecter_serie(
                ws2,
                r2
            )

            if not serie_cible.get("present"):
                continue

            ligne_dates = trouver_entete_dates(
                ws2,
                r2
            )


            warning = None
            if not ligne_dates:
                warning = "serie_detectee_sans_entete_temporel"
                
            dates = {}
            if ligne_dates:
                dates = lire_entete_dates(
                    ws2,
                    ligne_dates
                )

            segments = segmenter(
                ws2,
                r2,
                dates
            )

            # --------------------------------------------------
            # Succès
            # --------------------------------------------------

            return {
                "cellule_libelle": adresse_libelle,
                "cellule_source": adresse_cible,

                "libelle": libelle,
                "libelle_resolu": option,

                "structure": [
                    "selection",
                    "serie"
                ],

                "nature": nature,

                "detail": {
                    "resolution": "selector_to_series",

                    "selector_value": indice,
                    "selector_label": option,

                    "selector_option_address": option_info["address"],
                    "selector_option_distance" : option_info["distance"],

                    "source_series": adresse_cible,

                    "warning" : warning,
                    "ligne_dates" : ligne_dates,

                    "segments": segments,

                    "restitution": restituer(
                        segments,
                        nature
                    ),
                   
                }
            }

    # ==========================================================
    # Aucun candidat du catalogue n'a donné de série
    # ==========================================================

    return {
        "cellule_libelle": adresse_libelle,
        "libelle": libelle,
        "structure": ["selection"],
        "nature": nature,

        "detail": {
            "resolution": "selector_target_not_found",
            "selector_value": indice,
            "selector_options": options,

            "reason": (
                "Le sélecteur a été compris, mais aucune "
                "série correspondante n'a été trouvée "
                "dans le catalogue."
            )
        }
    }

def resoudre(
    wb,
    adresse_libelle,
    adresse_valeur=None,
    nature=None,
    resolver=None,
    catalogue=None,
):
    """
    Point d'entrée unique de résolution.

    Principe :
      - resolvers simples -> utilisent directement adresse_valeur ;
      - resolvers spécialisés -> traitement spécifique ;
      - fallback générique -> uniquement si aucune stratégie
        appropriée n'est déclarée.
    """

    # ----------------------------------------------------------
    # 1. Résolution directe des métriques scalaires
    # ----------------------------------------------------------

    RESOLVERS_DIRECTS = {
        "scalar_value",
        "rate_value",
        "year1_generation",
        "scalar_result",
        "duration_value",
    }

    if resolver in RESOLVERS_DIRECTS:
        return resoudre_valeur_directe(
            wb=wb,
            adresse_libelle=adresse_libelle,
            adresse_valeur=adresse_valeur,
            nature=nature,
        )

    # ----------------------------------------------------------
    # 2. Resolver spécialisé : sélecteur -> série
    # ----------------------------------------------------------

    if resolver == "index_selector_to_series":
        return resoudre_index_selector_to_series(
            wb=wb,
            adresse_libelle=adresse_libelle,
            adresse_valeur=adresse_valeur,
            nature=nature,
            catalogue=catalogue,
        )

    # ----------------------------------------------------------
    # 3. Fallback générique
    # ----------------------------------------------------------

    return _resoudre_generique(
        wb,
        adresse_libelle,
        nature=nature,
    )


def resumer(resultat):
    """
    Rend une ligne de synthèse lisible pour l'analyste.
    Compatible avec :
      - résolution générique
      - série directe
      - sélecteur résolu vers série
      - sélecteur non résolu
    """

    s = resultat.get("structure", [])
    d = resultat.get("detail", {})

    # ----------------------------------------------------------
    # 1. TOTAL
    # ----------------------------------------------------------
    if d.get("total"):
        t = d["total"]

        return (
            f"{t['adresse']} = "
            f"{t['valeur']:,.2f}"
        ).rstrip("0").rstrip(".") + " (total)"

    # ----------------------------------------------------------
    # 2. SÉRIE
    #    directe OU obtenue après résolution d'un sélecteur
    # ----------------------------------------------------------
    if d.get("restitution"):

        restitution = d.get("restitution") or []
        txt = " | ".join(restitution)

        # série obtenue via un sélecteur
        if d.get("resolution") == "selector_to_series":

            label = d.get("selector_label", "?")
            source = d.get("source_series")

            if source:
                return (
                    f"{label} [{source}] → "
                    f"{txt or '(série vide)'}"
                )

            return (
                f"{label} → "
                f"{txt or '(série vide)'}"
            )

        # série trouvée directement
        if d.get("resolution") == "direct_series":
            return txt or "(série vide)"

        # ancienne logique série
        return txt or "(série vide)"

    # ----------------------------------------------------------
    # 3. VALEUR SIMPLE
    # ----------------------------------------------------------
    if "valeur" in s and d.get("valeur") is not None:

        adresse = d.get("adresse_valeur", "?")
        valeur = d.get("valeur")

        return f"{adresse} = {valeur}"

    # ----------------------------------------------------------
    # 4. ANCIEN FORMAT DE SÉLECTION
    # ----------------------------------------------------------
    if "selection" in s and "selecteur" in d:

        return (
            f"sélecteur {d.get('selecteur')} "
            f"= option {d.get('option_active')}"
        )

    # ----------------------------------------------------------
    # 5. NOUVEAU FORMAT — SÉLECTEUR NON RÉSOLU
    # ----------------------------------------------------------
    if "selection" in s and d.get("selector_value") is not None:

        indice = d.get("selector_value")
        resolution = d.get("resolution", "non_resolu")

        return (
            f"sélecteur = {indice} "
            f"({resolution})"
        )

    # ----------------------------------------------------------
    # 6. ÉCHEC EXPLICITE
    # ----------------------------------------------------------
    if d.get("reason"):

        return f"non résolu : {d['reason']}"

    # ----------------------------------------------------------
    # 7. CAS INCONNU
    # ----------------------------------------------------------
    return "(structure inconnue)"
