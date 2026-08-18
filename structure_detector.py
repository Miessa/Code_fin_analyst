# -*- coding: utf-8 -*-
"""
structure_detector.py — OBSERVE des faits sur un candidat (ne juge pas).

Deux responsabilités séparées :
  • observer()   : lit des FAITS bruts sur la cellule (type Excel, format, formule,
                   en-tête temporel, mots-clés structurels, unité). AUCUNE déduction.
  • classifier() : à partir des faits, DÉDUIT une structure (scalar / scalar_aggregate
                   / time_series / selector_candidate). Modifiable sans toucher à la
                   lecture Excel.

Corrections appliquées (revue) :
  - le détecteur REÇOIT (adresse_libelle, adresse_valeur) : il ne cherche plus
    lui-même « la première valeur à droite » (fragile) ;
  - value_type = type Excel PHYSIQUE (numeric/text/date/boolean), PAS déduit de la
    magnitude ; + number_format_hint lu du format Excel (fait fiable) ;
  - unit_family = unknown par défaut ; currency_amount seulement si token monétaire ;
  - énergie testée AVANT puissance (mwh contient mw) ;
  - série = valeurs alignées SOUS un en-tête temporel (dates, années, Q1.., Year N) ;
  - has_total_keyword (lexical) séparé de is_aggregate_formula (=SUM...) ;
  - sélecteur : has_selector_keyword + is_small_integer (pas de plage 1..8 figée) ;
  - decomposition retirée (non détectée de façon fiable pour l'instant).
"""
import re, datetime
from openpyxl.utils import get_column_letter, column_index_from_string

MOTS_SELECTEUR = ("selector", "scenario", "scénario", "switch", "selection", "choice")
MOTS_TOTAL = ("total", "sum", "aggregate", "grand total", "sous-total", "subtotal")
TOKENS_MONNAIE = ("eur", "usd", "xaf", "fcfa", "cfa", "€", "$", "k€", "'000", "million", "millions", "mrd", "md")
FONCTIONS_AGREGAT = ("SUM", "SUBTOTAL", "SUMIF", "SUMIFS", "AGGREGATE")
SEUIL_SERIE = 5


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _coord(adresse):
    feuille, cell = adresse.split("!", 1)
    col = "".join(ch for ch in cell if ch.isalpha())
    row = int("".join(ch for ch in cell if ch.isdigit()))
    return feuille, column_index_from_string(col), row


def _excel_type(v):
    if v is None:
        return "empty"
    if isinstance(v, bool):
        return "boolean"
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "date"
    if isinstance(v, (int, float)):
        return "numeric"
    return "text"


def _format_hint(number_format):
    """Déduit un indice de nature à partir du FORMAT Excel (fait fiable)."""
    if not number_format or number_format == "General":
        return "general"
    nf = number_format.lower()
    if "%" in nf:
        return "percentage"
    if any(t in nf for t in ("€", "$", "eur", "usd", "xaf", "fcfa", "[$")):
        return "currency"
    if any(t in nf for t in ("yy", "mm", "dd", "yyyy")):
        return "date"
    return "general"


def _unit_family(libelle, number_format_hint):
    """unknown par défaut. currency seulement si token monétaire ou format currency."""
    lib = (libelle or "").lower()
    if number_format_hint ==  "percentage" or "percentage" in lib or "%" in lib:
        return "ratio"
    if any(u in lib for u in ("gwh", "mwh","kwh","electricity generation","energy generation","production d'énergie",
        "productible", )):
        return "energy"
    if any(u in lib for u in ( "mw", "kw", "gw", "installed capacity", "plant capacity", "rated capacity",)):
        return "power"
    if any(u in lib for u in ( "duration", "durée", "tenor", "maturity", "construction period", "concession period",
        "useful life",)):
        return "duration"
    if any(u in lib for u in ("eur","usd","xaf","fcfa","cfa","cost","capex","opex","price","revenue",
    )):
        return "currency_amount"
    return "unknown"

def _temporal_scope(libelle):
    """
    Décrit QUAND la métrique s'applique,
    indépendamment de son unité.
    """

    txt = (libelle or "").lower()

    if any(x in txt for x in (
        "year 1",
        "year1",
        "first year",
        "année 1",
        "annee 1",
    )):
        return "year_1"

    if any(x in txt for x in (
        "at cod",
        "commercial operation date",
    )):
        return "cod"

    if any(x in txt for x in (
        "during construction",
        "construction phase",
    )):
        return "construction"

    if any(x in txt for x in (
        "during operations",
        "operations",
        "operating period",
    )):
        return "operations"

    if any(x in txt for x in (
        "lifetime",
        "project life",
    )):
        return "lifetime"

    if any(x in txt for x in (
        "annual",
        "yearly",
        "per year",
        "annuel",
    )):
        return "annual"

    if any(x in txt for x in (
        "monthly",
        "per month",
        "mensuel",
    )):
        return "monthly"

    return "unknown"

def _entete_temporel(ws_cache, ws, ligne_valeur, ci, cf, remonter=250):
    """Cherche AU-DESSUS une ligne d'en-tête temporel : dates, années (2024..2075),
    trimestres (Q1..), mois (Jan..), ou 'Year N'. Retourne (trouvé, ligne)."""
    an = re.compile(r"^(19|20)\d{2}$")
    mois = ("jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec",
            "janv", "févr", "mars", "avr", "mai", "juin", "juil", "août", "sept", "oct", "nov", "déc")
    for rr in range(ligne_valeur - 1, max(0, ligne_valeur - remonter), -1):
        vr = _row(ws_cache, ws, rr)
        n_dates = n_annees = n_qm = 0
        for x in vr[ci-1:cf]:
            if isinstance(x, (datetime.datetime, datetime.date)):
                n_dates += 1
            elif isinstance(x, (int, float)) and not isinstance(x, bool) and an.match(str(int(x))):
                n_annees += 1
            elif isinstance(x, str):
                xs = x.strip().lower()
                if xs.startswith(("q1", "q2", "q3", "q4", "year", "yr", "an ", "année")) or xs[:3] in mois:
                    n_qm += 1
        if max(n_dates, n_annees, n_qm) >= 5:
            return True, rr
    return False, None


def _row(cache, ws, r):
    if r in cache:
        return cache[r]
    try:
        vals = [c.value for c in next(ws.iter_rows(min_row=r, max_row=r))]
    except StopIteration:
        vals = []
    cache[r] = vals
    return vals


def observer(wb, adresse_libelle, adresse_valeur, wb_formules=None, c1="J", c2="HA"):
    """Observe les FAITS bruts d'un candidat (label + valeur fournis par le collecteur)."""
    fl, cl, rl = _coord(adresse_libelle)
    fv, cv, rv = _coord(adresse_valeur)
    ws = wb[fl]
    cache = {}

    libelle = None
    ligne_lib = _row(cache, ws, rl)
    if cl - 1 < len(ligne_lib):
        libelle = ligne_lib[cl - 1]
    lib_l = libelle.lower() if isinstance(libelle, str) else ""

    # valeur + type Excel physique
    wsv = wb[fv]
    cell_v = wsv.cell(rv, cv)
    valeur = cell_v.value
    excel_type = _excel_type(valeur)
    number_format = getattr(cell_v, "number_format", None)
    number_format_hint = _format_hint(number_format)

    # mots-clés structurels (observations lexicales)
    has_total_keyword = any(m in lib_l for m in MOTS_TOTAL)
    has_selector_keyword = any(m in lib_l for m in MOTS_SELECTEUR)
    is_small_integer = _num(valeur) and float(valeur).is_integer() and abs(valeur) < 100

    # formule (si classeur formules fourni)
    is_formula = False; formula = None; formula_function = None; is_aggregate_formula = False
    if wb_formules is not None:
        try:
            fcell = wb_formules[fv].cell(rv, cv).value
            if isinstance(fcell, str) and fcell.startswith("="):
                is_formula = True; formula = fcell
                m = re.match(r"=\s*([A-Z]+)\s*\(", fcell.upper())
                if m:
                    formula_function = m.group(1)
                    is_aggregate_formula = formula_function in FONCTIONS_AGREGAT
        except Exception:
            pass

    # en-tête temporel + valeurs alignées dessous (sur la LIGNE du libellé)
    ci, cf = column_index_from_string(c1), column_index_from_string(c2)
    has_time_header, ligne_header = _entete_temporel(cache, ws, rl, ci, cf)
    ligne_vals = _row(cache, ws, rl)
    aligned_numeric_count = sum(1 for x in ligne_vals[ci-1:cf] if _num(x))

    unit_family = _unit_family(libelle, number_format_hint)
    temporal_scope = _temporal_scope(libelle)

    return {
        "label": libelle if isinstance(libelle, str) else None,
        "label_address": adresse_libelle,
        "value_address": adresse_valeur,
        "value": valeur,
        "excel_type": excel_type,
        "number_format": number_format,
        "number_format_hint": number_format_hint,
        "has_total_keyword": has_total_keyword,
        "has_selector_keyword": has_selector_keyword,
        "is_small_integer": is_small_integer,
        "has_time_header": has_time_header,
        "aligned_numeric_count": aligned_numeric_count,
        "is_formula": is_formula,
        "formula": formula,
        "formula_function": formula_function,
        "is_aggregate_formula": is_aggregate_formula,
        "unit_family": unit_family,
        "temporal_scope": temporal_scope,
        
    }


def classifier(faits):
    """DÉDUIT une structure à partir des faits. Séparé de l'observation :
    si demain SUM ≠ agrégat, on ne change que cette fonction."""
    # série : valeurs alignées SOUS un en-tête temporel
    if faits.get("has_time_header") and faits.get("aligned_numeric_count", 0) >= SEUIL_SERIE \
       and not faits.get("has_total_keyword"):
        structure = "time_series"
    # agrégat : mot 'total' OU formule d'agrégation
    elif faits.get("has_total_keyword") or faits.get("is_aggregate_formula"):
        structure = "scalar_aggregate"
    # sélecteur candidat : mot sélecteur + petit entier
    elif faits.get("has_selector_keyword") and faits.get("is_small_integer"):
        structure = "selector_candidate"
    else:
        structure = "scalar"

    # nature : d'après le FORMAT Excel, pas la magnitude
    hint = faits.get("number_format_hint")
    if hint == "percentage":
        value_type = "rate"
    elif faits.get("excel_type") == "numeric":
        value_type = "amount"
    elif faits.get("excel_type") in ("text", "date", "boolean"):
        value_type = faits["excel_type"]
    else:
        value_type = "unknown"

    # is_total : trois états.
    #  True  = preuve (mot 'total' OU formule d'agrégation)
    #  False = libellé lisible, aucun indice de total
    #  None  = information insuffisante (pas de libellé exploitable)
    a_libelle = bool(faits.get("label"))
    if faits.get("has_total_keyword") or faits.get("is_aggregate_formula"):
        is_total = True
    elif a_libelle:
        is_total = False
    else:
        is_total = None

    is_selector = True if (faits.get("has_selector_keyword") and faits.get("is_small_integer")) else False

    return {
        "structure": structure,
        "value_type": value_type,
        "unit_family": faits.get("unit_family", "unknown"),
        "temporal_scope": faits.get("temporal_scope", "unknown"),
        "is_total": is_total,
        "is_selector": is_selector,
        "is_time_series": structure == "time_series",
        # on propage les faits utiles au filtre / à l'analyste
        "value": faits.get("value"),
        "value_address": faits.get("value_address"),
        "label": faits.get("label"),
        "number_format_hint": faits.get("number_format_hint"),
    }



def caracteriser(wb, adresse_libelle, adresse_valeur, wb_formules=None):
    """Enchaîne observer -> classifier. Retourne le diagnostic complet."""
    faits = observer(wb, adresse_libelle, adresse_valeur, wb_formules)
    diag = classifier(faits)
    diag["_faits"] = faits
    return diag
