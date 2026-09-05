"""Deterministic adapter for the official IRENA RPGC Excel datafile."""

from __future__ import annotations

import hashlib
import re
import unicodedata
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter
from pydantic import Field

from benchmark_bank.schemas import BenchmarkBankDocument, IngestionRun, NormalizationEvent, Observation, Source
from benchmark_bank.schemas.common import CanonicalModel


SOURCE_ID = "irena_rpgc_2024"
PUBLICATION_URL = "https://www.irena.org/Publications/2025/Jun/Renewable-Power-Generation-Costs-in-2024"
DATA_URL = "https://www.irena.org/-/media/Files/IRENA/Agency/Publication/2025/Jul/IRENA-Datafile-RenPwrGenCosts-in-2024.xlsx"
TECH_MAP = {"Bioenergy":"bioenergy", "Geothermal":"geothermal", "Hydropower":"hydropower",
            "Hydro":"hydropower", "Solar PV":"solar_pv", "Solar photovoltaic":"solar_pv",
            "CSP":"solar_csp", "Solar thermal":"solar_csp", "Onshore wind":"onshore_wind",
            "Offshore wind":"offshore_wind", "Wind power":"wind_power", "Biomass for power":"bioenergy"}


class IRENATabularReport(CanonicalModel):
    source_id: str = SOURCE_ID
    artifact_path: str
    artifact_checksum_sha256: str
    sheets_read: list[str] = Field(default_factory=list)
    observation_count: int
    observations_by_metric: dict[str, int] = Field(default_factory=dict)
    observations_by_statistic: dict[str, int] = Field(default_factory=dict)
    extraction_method: str = "deterministic_xlsx"
    llm_calls: int = 0
    warnings: list[str] = Field(default_factory=list)


def _sha256(path):
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""): digest.update(block)
    return digest.hexdigest()


def _number(value):
    if value is None or str(value).strip().lower() in {"n.a", "n.a.", "na", ""}: return None
    try: return float(str(value).replace(" ", "").replace("\u00a0", ""))
    except (TypeError, ValueError): return None


def _token(value):
    """Create stable schema-safe identifiers from arbitrary labels."""
    text = unicodedata.normalize("NFKD", str(value)).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return text[:48] or hashlib.sha256(str(value).encode("utf-8")).hexdigest()[:12]


class IRENATabularAdapter:
    adapter_name, adapter_version = "irena_rpgc_tabular", "1.0.0"
    def __init__(self, artifact_path): self.artifact_path = Path(artifact_path)
    def read(self): return openpyxl.load_workbook(self.artifact_path, read_only=True, data_only=True)
    def build(self): return self.build_from_workbook(self.read(), checksum=_sha256(self.artifact_path))

    def build_from_workbook(self, workbook, checksum="0" * 64):
        required = {"Table S.1", "Fig 1.2 ", "Table 6.2", "Table 6.3", "Table 6.4", "Table 6.5",
                    "Table 6.6", "Table A1", "Fig A1", "Table A4", "Table A5"}
        missing = sorted(required - set(workbook.sheetnames))
        if missing: raise ValueError(f"IRENA workbook sheets missing: {missing}")
        run_id = f"irena:rpgc2024:{checksum[:16]}"
        source = Source(source_id=SOURCE_ID, organization="International Renewable Energy Agency (IRENA)",
            title="Renewable Power Generation Costs in 2024 — Datafile", source_type="sector_cost_dataset",
            publication_date=date(2025,7,1), data_period_start=date(2010,1,1), data_period_end=date(2024,12,31),
            canonical_url=DATA_URL, content_checksum_sha256=checksum, review_status="pending",
            notes="Official tabular data accompanying the report; sector statistics, not identified projects.")
        observations, events, counts, stat_counts, warnings = [], [], {}, {}, []

        def add(identifier, metric, technology, value, unit, statistic, sheet, cell, *, low=None, high=None,
                geography="global", metadata=None, value_year=2024, perimeter="newly commissioned utility-scale projects"):
            if value is None and low is None: return
            oid = f"irena:{identifier}"
            observations.append(Observation(observation_id=oid, source_id=SOURCE_ID, project_id=None,
                ingestion_run_id=run_id, metric=metric, observation_type="sector_statistic",
                raw_value=value, raw_low=low, raw_high=high, raw_unit=unit,
                normalized_value=value, normalized_low=low, normalized_high=high, normalized_unit=unit,
                currency="USD" if "USD" in unit else None, price_year=2024 if "USD" in unit else None,
                value_date=date(value_year,12,31), statistic=statistic, economic_perimeter=perimeter,
                source_location=f"Workbook sheet '{sheet}', {cell}", quality_level="high", review_status="pending",
                metadata={"technology":technology, "geography":geography, "method":"published_tabular", **(metadata or {})}))
            events.append(NormalizationEvent(normalization_event_id=f"norm:{oid}", ingestion_run_id=run_id,
                observation_id=oid, field_name="normalized_value", rule_id="rule:tabular_canonical_identity",
                rule_version="1.0.0", input_value=value if value is not None else [low,high],
                output_value=value if value is not None else [low,high]))
            counts[metric] = counts.get(metric,0)+1; stat_counts[statistic] = stat_counts.get(statistic,0)+1

        # Global 2024 headline metrics by technology.
        ws = workbook["Table S.1"]
        for row in range(7, ws.max_row + 1):
            label, tech = ws.cell(row,2).value, TECH_MAP.get(str(ws.cell(row,2).value).strip())
            if not tech: continue
            add(f"{tech}:tic:2024", "total_installed_cost", tech, _number(ws.cell(row,4).value),
                "2024 USD/kW", "weighted_average", "Table S.1", f"D{row}")
            capacity_factor = _number(ws.cell(row,7).value)
            add(f"{tech}:capacity_factor:2024", "capacity_factor", tech,
                capacity_factor / 100 if capacity_factor is not None else None,
                "ratio", "weighted_average", "Table S.1", f"G{row}")
            add(f"{tech}:lcoe:2024", "lcoe", tech, _number(ws.cell(row,10).value),
                "2024 USD/kWh", "weighted_average", "Table S.1", f"J{row}")

        # Hydropower installed-cost distribution by capacity band, 2010-2024.
        ws = workbook["Table 6.2"]
        for row in range(5, ws.max_row + 1):
            band = ws.cell(row,2).value
            low, avg, high = map(_number, (ws.cell(row,3).value, ws.cell(row,4).value, ws.cell(row,5).value))
            if band and avg is not None:
                safe = str(band).replace(" ","").replace(">", "gt").replace("<", "lt").replace("–","-")
                add(f"hydropower:tic:capacity_band:{safe}", "total_installed_cost", "hydropower", avg,
                    "2024 USD/kW", "weighted_average_with_p05_p95", "Table 6.2", f"B{row}:E{row}", low=low, high=high,
                    metadata={"capacity_band_mw":str(band), "period_start":2010, "period_end":2024})

        # Hydropower regional TIC by size and period.
        ws = workbook["Table 6.3"]
        for row in range(6, ws.max_row + 1):
            geography = ws.cell(row,3).value
            if not geography: continue
            for col, size, period in ((4,"large","2010_2017"),(5,"large","2018_2024"),(6,"small","2010_2017"),(7,"small","2018_2024")):
                value = _number(ws.cell(row,col).value)
                add(f"hydropower:tic:{_token(geography)}:{size}:{period}",
                    "total_installed_cost", "hydropower", value, "2024 USD/kW", "weighted_average",
                    "Table 6.3", f"{get_column_letter(col)}{row}", geography=str(geography),
                    metadata={"hydropower_size":size, "period":period})

        # Regional capacity-factor distributions for large and small hydro.
        for sheet, size in (("Table 6.4","large"),("Table 6.5","small")):
            ws = workbook[sheet]
            for row in range(7, ws.max_row + 1):
                geography = ws.cell(row,2).value
                if not geography: continue
                for start_col, period in ((3,"2010_2017"),(6,"2018_2024")):
                    low, avg, high = [_number(ws.cell(row,c).value) for c in range(start_col,start_col+3)]
                    if avg is not None:
                        add(f"hydropower:cf:{_token(geography)}:{size}:{period}",
                            "capacity_factor", "hydropower", avg/100, "ratio", "weighted_average_with_p05_p95",
                            sheet, f"{get_column_letter(start_col)}{row}:{get_column_letter(start_col + 2)}{row}",
                            low=low/100 if low is not None else None, high=high/100 if high is not None else None,
                            geography=str(geography), metadata={"hydropower_size":size,"period":period})

        # Historical global LCOE series, preserving every published year.
        ws = workbook["Fig 1.2 "]
        years = [(col, int(ws.cell(5, col).value)) for col in range(3, ws.max_column + 1)
                 if _number(ws.cell(5, col).value) is not None]
        for row in range(6, 13):
            tech = TECH_MAP.get(str(ws.cell(row, 2).value).strip())
            if not tech: continue
            for col, year in years:
                if year == 2024: continue  # already captured from the headline table
                add(f"{tech}:lcoe:{year}", "lcoe", tech, _number(ws.cell(row, col).value),
                    "2024 USD/kWh", "weighted_average", "Fig 1.2 ",
                    f"{get_column_letter(col)}{row}", value_year=year)

        # Distribution of hydropower O&M cost components across 25 projects.
        ws = workbook["Table 6.6"]
        for row in range(5, ws.max_row + 1):
            component = ws.cell(row, 2).value
            average = _number(ws.cell(row, 4).value)
            if component and average is not None:
                add(f"hydropower:opex_component:{_token(component)}", "opex_component_share", "hydropower",
                    average / 100, "ratio", "weighted_average_with_min_max", "Table 6.6", f"B{row}:E{row}",
                    low=(_number(ws.cell(row, 3).value) or 0) / 100,
                    high=(_number(ws.cell(row, 5).value) or 0) / 100,
                    metadata={"component": str(component), "sample_size": 25},
                    perimeter="share of total hydropower O&M costs")

        # Standard economic-life assumptions.
        ws = workbook["Table A1"]
        for row in range(6, ws.max_row + 1):
            label, life = ws.cell(row,2).value, _number(ws.cell(row,3).value)
            tech = TECH_MAP.get(str(label).strip())
            if tech and life is not None:
                add(f"{tech}:economic_life", "economic_life", tech, life, "year", "lcoe_assumption",
                    "Table A1", f"C{row}", perimeter="standardised LCOE calculation assumption")

        # Country- and technology-specific real after-tax WACC assumptions.
        ws = workbook["Fig A1"]
        headers = [(col, TECH_MAP.get(str(ws.cell(3,col).value).strip())) for col in range(3,10)]
        for row in range(4, ws.max_row + 1):
            country = ws.cell(row,2).value
            if not country: continue
            for col, tech in headers:
                value = _number(ws.cell(row,col).value)
                if tech and value is not None:
                    add(f"{tech}:wacc:{_token(country)}:2024", "wacc", tech,
                        value, "ratio", "model_assumption", "Fig A1", f"{get_column_letter(col)}{row}",
                        geography=str(country), perimeter="real after-tax WACC assumption")

        # Fixed O&M assumptions used by IRENA's LCOE calculations.
        ws = workbook["Table A4"]
        for row in range(5, ws.max_row + 1):
            year = _number(ws.cell(row, 2).value)
            if year is None: continue
            for col, geography in ((3, "OECD"), (4, "Non-OECD")):
                add(f"solar_pv:opex_fixed:{_token(geography)}:{int(year)}", "opex_fixed", "solar_pv",
                    _number(ws.cell(row, col).value), "2024 USD/kW/year", "lcoe_assumption", "Table A4",
                    f"{get_column_letter(col)}{row}", geography=geography, value_year=int(year),
                    perimeter="standardised LCOE calculation assumption")
        ws = workbook["Table A5"]
        for row in range(4, ws.max_row + 1):
            geography = ws.cell(row, 2).value
            if geography:
                add(f"offshore_wind:opex_fixed:{_token(geography)}:2024", "opex_fixed", "offshore_wind",
                    _number(ws.cell(row, 3).value), "2024 USD/kW/year", "lcoe_assumption", "Table A5",
                    f"C{row}", geography=str(geography), perimeter="standardised LCOE calculation assumption")

        if len(observations) < 100: raise ValueError("IRENA tabular extraction produced unexpectedly few observations")
        run = IngestionRun(ingestion_run_id=run_id, source_id=SOURCE_ID, adapter_name=self.adapter_name,
            adapter_version=self.adapter_version, status="succeeded", started_at=datetime.now(timezone.utc),
            completed_at=datetime.now(timezone.utc), raw_artifact_path=str(self.artifact_path),
            raw_artifact_checksum_sha256=checksum, records_read=len(observations), records_accepted=len(observations),
            records_rejected=0, configuration={"report_year":2024,"llm_enabled":False,"source_format":"xlsx"})
        document = BenchmarkBankDocument(sources=[source],projects=[],ingestion_runs=[run],observations=observations,
                                         normalization_events=events)
        report = IRENATabularReport(artifact_path=str(self.artifact_path),artifact_checksum_sha256=checksum,
            sheets_read=sorted(required),observation_count=len(observations),observations_by_metric=counts,
            observations_by_statistic=stat_counts,warnings=warnings)
        return document, report
