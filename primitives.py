# -*- coding: utf-8 -*-
"""
primitives.py — DÉTECTEURS DE PRIMITIVES DE STRUCTURE

Un modèle de promoteur structure ses métriques de façons variées. Plutôt que
d'énumérer les structures (liste ouverte, impossible), on reconnaît 4 PRIMITIVES
élémentaires qui se combinent :

  • VALEUR SIMPLE   — un scalaire
  • SÉLECTION       — un choix parmi options (sélecteur 1..n)
  • SÉRIE TEMPORELLE— une valeur qui varie sur un axe de dates
  • DÉCOMPOSITION   — un poste éclaté en composantes (+ un total)

Chaque détecteur est DÉTERMINISTE (le code décide) et répond :
  {"present": bool, ...données extraites...}
L'outil les applique en cascade et les COMPOSE. La structure n'est pas
déclarée d'avance : elle est détectée à l'exécution, fichier par fichier.

Le LLM n'intervient pas ici : il ne sert qu'ensuite, pour RÉSOUDRE le sens
(ex. « ce sélecteur=2 correspond à quel indice ? »), jamais pour détecter.
"""
import re
import datetime
from openpyxl.utils import get_column_letter, column_index_from_string

# ---- Réglages ------------------------------------------------------------
MIN_POINTS_SERIE = 5        # nb de valeurs alignées pour parler de série
PLAGE_SELECTEUR = (1, 8)    # un sélecteur est un entier dans cette plage
PORTEE_VOISINE = 10         # combien de colonnes à droite on inspecte
MOTS_SELECTEUR = ("selector", "scenario", "scénario", "select", "switch", "choice", "flag")
# -------------------------------------------------------------------------


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ═══════════════════════════════════════════════════════════════════════
# PRIMITIVE 1 — VALEUR SIMPLE
# ═══════════════════════════════════════════════════════════════════════
def detecter_valeur(ws, r, c, portee=PORTEE_VOISINE):
    """Cherche un scalaire voisin (droite puis dessous). Toujours le dernier recours."""
    for cc in range(c + 1, c + portee + 1):
        v = ws.cell(r, cc).value
        if _num(v):
            return {"present": True, "type": "valeur", "valeur": v,
                    "adresse": f"{ws.title}!{get_column_letter(cc)}{r}"}
    for rr in range(r + 1, r + 3):
        v = ws.cell(rr, c).value
        if _num(v):
            return {"present": True, "type": "valeur", "valeur": v,
                    "adresse": f"{ws.title}!{get_column_letter(c)}{rr}"}
    return {"present": False, "type": "valeur"}


# ═══════════════════════════════════════════════════════════════════════
# PRIMITIVE 2 — SÉLECTION (un sélecteur pilote un choix)
# ═══════════════════════════════════════════════════════════════════════
def detecter_selection(ws, r, c, portee=PORTEE_VOISINE):
    """Repère un sélecteur : soit le libellé courant contient 'selector/scenario'
    et a une valeur entière voisine, soit une ligne voisine en a un.
    Retourne aussi les options (colonnes de scénario J..O si présentes)."""
    lo, hi = PLAGE_SELECTEUR
    # a) le libellé lui-même est un sélecteur
    lib = ws.cell(r, c).value
    est_sel_lib = isinstance(lib, str) and any(m in lib.lower() for m in MOTS_SELECTEUR)
    for cc in range(c + 1, c + portee + 1):
        v = ws.cell(r, cc).value
        if _num(v) and lo <= v <= hi and float(v).is_integer():
            if est_sel_lib:
                return {"present": True, "type": "selection",
                        "selecteur": f"{ws.title}!{get_column_letter(cc)}{r}",
                        "valeur_active": int(v),
                        "libelle": lib.strip() if isinstance(lib, str) else None}
    # b) un sélecteur sur la ligne juste au-dessus/dessous (pilote le bloc)
    for rr in (r - 1, r + 1):
        libr = ws.cell(rr, c).value
        if isinstance(libr, str) and any(m in libr.lower() for m in MOTS_SELECTEUR):
            for cc in range(c + 1, c + portee + 1):
                v = ws.cell(rr, cc).value
                if _num(v) and lo <= v <= hi and float(v).is_integer():
                    return {"present": True, "type": "selection",
                            "selecteur": f"{ws.title}!{get_column_letter(cc)}{rr}",
                            "valeur_active": int(v), "libelle": libr.strip()}
    return {"present": False, "type": "selection"}


# ═══════════════════════════════════════════════════════════════════════
# PRIMITIVE 3 — SÉRIE TEMPORELLE (valeurs alignées sur un axe de dates)
# ═══════════════════════════════════════════════════════════════════════
def detecter_serie(ws, r, c1="J", c2="HA", min_points=MIN_POINTS_SERIE):
    """Détecte une série : >= min_points valeurs numériques alignées sur la ligne."""
    ci, cf = column_index_from_string(c1), column_index_from_string(c2)
    vals = {c: ws.cell(r, c).value for c in range(ci, cf + 1) if _num(ws.cell(r, c).value)}
    if len(vals) >= min_points:
        return {"present": True, "type": "serie", "n_points": len(vals),
                "premiere_col": get_column_letter(min(vals)),
                "derniere_col": get_column_letter(max(vals))}
    return {"present": False, "type": "serie"}


def trouver_entete_dates(ws, ligne_serie, c1="J", c2="HA", remonter=200):
    """Cherche au-dessus de la série une ligne d'en-tête contenant des dates."""
    ci, cf = column_index_from_string(c1), column_index_from_string(c2)
    for rr in range(ligne_serie - 1, max(0, ligne_serie - remonter), -1):
        n = sum(1 for c in range(ci, cf + 1)
                if isinstance(ws.cell(rr, c).value, (datetime.datetime, datetime.date)))
        if n >= 5:
            return rr
    return None


# ═══════════════════════════════════════════════════════════════════════
# PRIMITIVE 4 — DÉCOMPOSITION (composantes + total)
# ═══════════════════════════════════════════════════════════════════════
def detecter_decomposition(ws, r, c, motif_composante=None, portee_bas=20):
    """Détecte si, sous le libellé courant, plusieurs lignes 'composante' existent,
    et si un 'Total ...' est présent à proximité. Heuristique déterministe :
    des libellés voisins partageant un préfixe commun + une ligne 'total'."""
    lib0 = ws.cell(r, c).value
    if not isinstance(lib0, str):
        return {"present": False, "type": "decomposition"}
    composantes = []
    total = None
    for rr in range(r, r + portee_bas):
        v = ws.cell(rr, c).value
        if not isinstance(v, str):
            continue
        vl = v.lower()
        if vl.startswith("total") or "total" in vl[:8]:
            # valeur du total
            for cc in range(c + 1, c + PORTEE_VOISINE + 1):
                tv = ws.cell(rr, cc).value
                if _num(tv):
                    total = {"libelle": v.strip(),
                             "adresse": f"{ws.title}!{get_column_letter(cc)}{rr}",
                             "valeur": tv}
                    break
        else:
            # composante = a une valeur numérique voisine
            for cc in range(c + 1, c + PORTEE_VOISINE + 1):
                cv = ws.cell(rr, cc).value
                if _num(cv):
                    composantes.append({"libelle": v.strip(),
                                        "adresse": f"{ws.title}!{get_column_letter(cc)}{rr}",
                                        "valeur": cv})
                    break
    # décomposition = au moins 2 composantes ET (idéalement) un total
    if len(composantes) >= 2:
        return {"present": True, "type": "decomposition",
                "composantes": composantes, "total": total}
    return {"present": False, "type": "decomposition"}


# ═══════════════════════════════════════════════════════════════════════
# CASCADE — compose les primitives pour caractériser UNE métrique
# ═══════════════════════════════════════════════════════════════════════
def caracteriser(ws, r, c):
    """Applique les détecteurs et retourne la structure composée détectée.
    Ordre : série → sélection → décomposition → valeur (recours).
    Les primitives se cumulent (une métrique peut être décomposée ET sélectionnée)."""
    profil = {"cellule_libelle": f"{ws.title}!{get_column_letter(c)}{r}",
              "libelle": ws.cell(r, c).value, "primitives": []}

    serie = detecter_serie(ws, r)
    if serie["present"]:
        serie["ligne_dates"] = trouver_entete_dates(ws, r)
        profil["primitives"].append(serie)

    selection = detecter_selection(ws, r, c)
    if selection["present"]:
        profil["primitives"].append(selection)

    decomp = detecter_decomposition(ws, r, c)
    if decomp["present"]:
        profil["primitives"].append(decomp)

    # valeur simple : seulement si rien d'autre n'a été détecté
    if not profil["primitives"]:
        val = detecter_valeur(ws, r, c)
        if val["present"]:
            profil["primitives"].append(val)

    profil["types"] = [p["type"] for p in profil["primitives"]] or ["inconnu"]
    return profil
