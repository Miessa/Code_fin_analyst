# -*- coding: utf-8 -*-
"""Benchmark Kikot : retrieval → structure → sélection finale."""

import argparse
import json
import time

from openpyxl import load_workbook

from arsel_core.workflow import doit_appeler_llm
from arsel_core.candidate_pipeline import recuperer_candidats, scorer_candidats
from arsel_core.collecter_libelles import collecter
from .evaluer_tfidf import charger_concepts, charger_json, normaliser_registre, _adresses_verite
from arsel_core.gemini_provider import appeler_json, disponible
from arsel_core.llm_instrumentation import InstrumentationLLM
from arsel_core.semantic_selector import (
    OUTCOME_AMBIGUOUS,
    OUTCOME_NO_MATCH,
    OUTCOME_SELECTED,
    STATUS_NOT_REQUIRED,
    choisir_semantiquement,
)
from arsel_core.tfidf_search import IndexTfidf
from arsel_core.embedding_search import creer_index_embeddings
from arsel_core.formula_dependency import IndexDependancesFormules


def _rang(candidats, adresses):
    return next((i for i, c in enumerate(candidats, 1)
                 if c.get("cellule_libelle") in adresses), None)


def _ratio(numerateur, denominateur):
    return numerateur / denominateur if denominateur else None


def construire_resume(details, instrumentation, duree):
    total = len(details)
    automatiques = [d for d in details if d["final_outcome"] == OUTCOME_SELECTED]
    corrects = [d for d in automatiques if d["final_correct"]]
    faux = [d for d in automatiques if not d["final_correct"]]
    fallback_corrects = sum(d["deterministic_fallback_correct"] for d in details)
    return {
        "nombre_evalue": total,
        "retrieval_recall@5": _ratio(sum(d["retrieval_rank"] is not None and d["retrieval_rank"] <= 5 for d in details), total),
        "retrieval_recall@10": _ratio(sum(d["retrieval_rank"] is not None and d["retrieval_rank"] <= 10 for d in details), total),
        "retrieval_recall@20": _ratio(sum(d["retrieval_rank"] is not None and d["retrieval_rank"] <= 20 for d in details), total),
        "structural_recall@5": _ratio(sum(d["structural_rank"] is not None and d["structural_rank"] <= 5 for d in details), total),
        "structural_top1_accuracy": _ratio(sum(d["structural_rank"] == 1 for d in details), total),
        "deterministic_fallback_accuracy": _ratio(fallback_corrects, total),
        "final_accuracy": _ratio(len(corrects), total),
        "automatic_selection_accuracy": _ratio(len(corrects), len(automatiques)),
        "false_automatic_selection_rate": _ratio(len(faux), len(automatiques)),
        "ambiguity_rate": _ratio(sum(d["final_outcome"] == OUTCOME_AMBIGUOUS for d in details), total),
        "no_match_rate": _ratio(sum(d["final_outcome"] == OUTCOME_NO_MATCH for d in details), total),
        "llm": instrumentation.resume(),
        "runtime_seconds": round(duree, 3),
    }


def benchmark(modele, referentiel, verite, avec_llm=False):
    debut = time.perf_counter()
    concepts = charger_concepts(referentiel)
    registre = normaliser_registre(charger_json(verite))
    catalogue = collecter(modele)
    index = IndexTfidf(catalogue)
    index_embeddings, erreur_embeddings = creer_index_embeddings(catalogue)
    # Le détecteur effectue de nombreux accès dispersés. Le mode read_only est
    # séquentiel et devient extrêmement lent ici ; le benchmark charge donc les
    # feuilles en mémoire pour conserver un temps d'exécution exploitable.
    wb = load_workbook(modele, data_only=True, read_only=False)
    wb_formules = load_workbook(modele, data_only=False, read_only=False)
    instrumentation = InstrumentationLLM(max_appels_par_metrique=1)
    index_dependances = IndexDependancesFormules(wb_formules)
    llm_actif = avec_llm and disponible()
    details = []

    for entree in registre:
        if entree.get("evaluation_enabled") is False:
            continue
        cle = entree.get("cle")
        concept = concepts.get(cle)
        attendues = _adresses_verite(entree)
        if not concept or not attendues:
            continue

        candidats, lexicaux, tfidf = recuperer_candidats(
            concept, catalogue, index,
            index_dependances=index_dependances,
            index_embeddings=index_embeddings,
        )
        scored = scorer_candidats(wb, concept, candidats, wb_formules=wb_formules)
        retrieval_rank = _rang(candidats, attendues)
        structural_rank = _rang(scored, attendues)
        top = scored[0] if scored else None
        appel_requis = bool(scored and doit_appeler_llm(scored))

        if not scored:
            decision = {"selection_outcome": OUTCOME_NO_MATCH,
                        "execution_status": STATUS_NOT_REQUIRED,
                        "cellule_libelle": None,
                        "raison": "Aucun candidat après retrieval."}
        elif not appel_requis:
            decision = {"selection_outcome": OUTCOME_SELECTED,
                        "execution_status": STATUS_NOT_REQUIRED,
                        "cellule_libelle": top["cellule_libelle"],
                        "raison": "Sélection structurelle déterministe."}
        elif llm_actif:
            decision = choisir_semantiquement(
                concept, scored[:8], instrumentation.instrumenter(cle, appeler_json)
            )
        else:
            decision = {"selection_outcome": OUTCOME_AMBIGUOUS,
                        "execution_status": STATUS_NOT_REQUIRED,
                        "cellule_libelle": None,
                        "raison": "LLM désactivé pour le benchmark."}

        adresse_finale = decision.get("cellule_libelle")
        details.append({
            "cle": cle,
            "ground_truth": sorted(attendues),
            "catalogue_present": any(a in {c.get("cellule_libelle") for c in catalogue} for a in attendues),
            "lexical_rank": _rang(lexicaux, attendues),
            "tfidf_rank": _rang(tfidf, attendues),
            "retrieval_rank": retrieval_rank,
            "structural_rank": structural_rank,
            "structural_top1": top.get("cellule_libelle") if top else None,
            "structural_top1_score": top.get("score") if top else None,
            "llm_required": appel_requis,
            "llm": instrumentation.pour_metrique(cle),
            "final_outcome": decision.get("selection_outcome"),
            "execution_status": decision.get("execution_status"),
            "final_address": adresse_finale,
            "deterministic_fallback": top.get("cellule_libelle") if top else None,
            "deterministic_fallback_correct": (
                top.get("cellule_libelle") in attendues if top else False
            ),
            "final_correct": adresse_finale in attendues if adresse_finale else False,
        })

    resume = construire_resume(details, instrumentation, time.perf_counter() - debut)
    return {"configuration": {"avec_llm": avec_llm, "llm_actif": llm_actif,
                               "catalogue_size": len(catalogue),
                               "embeddings_actifs": index_embeddings is not None,
                               "erreur_embeddings": erreur_embeddings},
            "summary": resume, "details": details}


def main(argv=None):
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("modele")
    parser.add_argument("verite")
    parser.add_argument("--referentiel", default="data/referentiel_arsel.json")
    parser.add_argument("--avec-llm", action="store_true")
    parser.add_argument("--sortie-json", default="evaluation/outputs/benchmark_pipeline_kikot.json")
    args = parser.parse_args(argv)
    rapport = benchmark(args.modele, args.referentiel, args.verite, args.avec_llm)
    with open(args.sortie_json, "w", encoding="utf-8") as fichier:
        json.dump(rapport, fichier, ensure_ascii=False, indent=2, default=str)
    print(json.dumps(rapport["summary"], ensure_ascii=False, indent=2))
    print(f"Rapport : {args.sortie_json}")


if __name__ == "__main__":
    main()
