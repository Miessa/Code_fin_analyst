import unittest

from openpyxl import Workbook

from benchmark_bank.sources import IRENATabularAdapter


def sample_workbook():
    workbook = Workbook()
    workbook.remove(workbook.active)
    sheets = ("Table S.1", "Fig 1.2 ", "Table 6.2", "Table 6.3", "Table 6.4", "Table 6.5",
              "Table 6.6", "Table A1", "Fig A1", "Table A4", "Table A5")
    for name in sheets:
        workbook.create_sheet(name)
    technologies = ("Bioenergy", "Geothermal", "Hydropower", "Solar PV", "CSP", "Onshore wind", "Offshore wind")
    summary = workbook["Table S.1"]
    for row, technology in enumerate(technologies, 7):
        summary.cell(row, 2, technology); summary.cell(row, 4, 2267 if technology == "Hydropower" else 1000 + row)
        summary.cell(row, 7, 45); summary.cell(row, 10, 0.06)
    costs = workbook["Table 6.2"]
    costs.cell(5, 2, "0-10 MW")
    for column, value in zip(range(3, 6), (1200, 2200, 4000)): costs.cell(5, column, value)
    regional = workbook["Table 6.3"]
    regional.cell(6, 3, "Africa & Middle East")
    for column in range(4, 8): regional.cell(6, column, 2400 + column)
    for sheet in ("Table 6.4", "Table 6.5"):
        capacity = workbook[sheet]; capacity.cell(7, 2, "Africa & Middle East")
        for column, value in zip(range(3, 9), (30, 45, 60, 32, 47, 62)): capacity.cell(7, column, value)
    life = workbook["Table A1"]
    for row, technology in enumerate(technologies, 6): life.cell(row, 2, technology); life.cell(row, 3, 30)
    wacc = workbook["Fig A1"]
    for column, technology in enumerate(technologies, 3): wacc.cell(3, column, technology)
    for row in range(4, 24):
        wacc.cell(row, 2, f"Country {row}")
        for column in range(3, 10): wacc.cell(row, column, 0.075)
    return workbook


class IRENATabularAdapterTests(unittest.TestCase):
    def test_extracts_tabular_statistics_without_projects_or_llm(self):
        document, report = IRENATabularAdapter("data.xlsx").build_from_workbook(sample_workbook())
        self.assertGreater(report.observation_count, 100)
        self.assertEqual(report.extraction_method, "deterministic_xlsx")
        self.assertEqual(report.llm_calls, 0)
        self.assertFalse(document.projects)
        hydro = next(item for item in document.observations if item.observation_id == "irena:hydropower:tic:2024")
        self.assertEqual(hydro.normalized_value, 2267)
        self.assertEqual(hydro.observation_type, "sector_statistic")
        self.assertIn("Table S.1", hydro.source_location)

    def test_fails_closed_when_required_sheet_is_missing(self):
        workbook = sample_workbook(); del workbook["Table A1"]
        with self.assertRaisesRegex(ValueError, "Table A1"):
            IRENATabularAdapter("data.xlsx").build_from_workbook(workbook)


if __name__ == "__main__": unittest.main()
