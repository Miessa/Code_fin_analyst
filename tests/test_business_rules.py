import unittest

from openpyxl import Workbook

from arsel_core.business_rules import evaluer_compatibilite
from arsel_core.resoudre import _inverser_segments, resoudre
from arsel_core.workflow import appliquer_unite_analyste


class TestBusinessRules(unittest.TestCase):
    def test_analyste_peut_corriger_unite_sans_changer_valeur(self):
        resultat = {
            "unite": "months",
            "resume": "Oper!F46 = 500 [months]",
            "detail": {"valeur": 500, "unite": "months"},
        }
        appliquer_unite_analyste(resultat, "MW")
        self.assertEqual(resultat["detail"]["valeur"], 500)
        self.assertEqual(resultat["unite"], "MW")
        self.assertEqual(resultat["resume"], "Oper!F46 = 500 [MW]")
        self.assertEqual(resultat["unite_source"], "analyste")

    def test_zero_est_toujours_aberrant(self):
        facteur, _, signaux = evaluer_compatibilite(
            {"cle": "wacc"}, {"libelle": "WACC", "valeur": 0}
        )
        self.assertEqual(facteur, 0.0)
        self.assertTrue(any("aberrante" in signal for signal in signaux))

    def test_duree_concession_rejette_une_grace_period_de_dette(self):
        facteur, _, signaux = evaluer_compatibilite(
            {"cle": "duree_concession"},
            {"libelle": "Senior debt replacement grace period - Tranche 4", "valeur": 6},
        )
        self.assertLess(facteur, 0.1)
        self.assertTrue(signaux)

    def test_productible_prefere_une_valeur_annuelle(self):
        annuel = evaluer_compatibilite(
            {"cle": "productible"},
            {"libelle": "Annual electricity generation", "valeur": 100},
        )[0]
        trimestriel = evaluer_compatibilite(
            {"cle": "productible"},
            {"libelle": "Electricity generation - 3 months", "valeur": 25},
        )[0]
        self.assertGreater(annuel, trimestriel)

    def test_inverse_unavailability(self):
        segments = [{"type": "palier", "valeur": 0.10, "de": "2025", "a": "2030"}]
        self.assertAlmostEqual(_inverser_segments(segments)[0]["valeur"], 0.90)

    def test_resolver_opex_year1_utilise_la_cellule_directe(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "InpC"
        ws["A1"] = "OPEX Year 1"
        ws["B1"] = 718
        resultat = resoudre(
            wb, "InpC!A1", "InpC!B1", nature="montant", resolver="year1_value"
        )
        self.assertEqual(resultat["detail"]["valeur"], 718)
        self.assertEqual(resultat["structure"], ["valeur"])


if __name__ == "__main__":
    unittest.main()
