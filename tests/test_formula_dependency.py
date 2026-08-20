import unittest

from openpyxl import Workbook

from arsel_core.formula_dependency import (
    IndexDependancesFormules,
    calculer_boost_dependances,
    concept_attend_agregat,
)


class TestFormulaDependency(unittest.TestCase):
    def setUp(self):
        self.wb = Workbook()
        ws = self.wb.active
        ws.title = "Summary"
        ws["A1"] = "=SUM(A2:A4)"
        ws["A2"], ws["A3"], ws["A4"] = 10, 20, 30
        ws["B1"] = "=A1"

    def test_detecte_agregat_et_dependances_recursives(self):
        analyse = IndexDependancesFormules(self.wb).analyser("Summary!B1")
        self.assertTrue(analyse["has_aggregate_formula"])
        self.assertGreaterEqual(analyse["dependency_count"], 4)
        self.assertEqual(analyse["leaf_count"], 3)

    def test_boost_exige_pertinence_lexicale(self):
        analyse = {"dependency_count": 10, "has_aggregate_formula": True}
        self.assertEqual(calculer_boost_dependances(analyse, 0.0), 0.0)
        self.assertGreater(calculer_boost_dependances(analyse, 0.1), 0.0)
        self.assertLessEqual(calculer_boost_dependances(analyse, 0.1), 0.2)

    def test_identifie_un_concept_agrege(self):
        self.assertTrue(concept_attend_agregat({"role_metrique": "aggregate"}))
        self.assertFalse(concept_attend_agregat({"role_metrique": "input"}))


if __name__ == "__main__": unittest.main()
