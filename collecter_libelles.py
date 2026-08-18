# -*- coding: utf-8 -*-
"""
collecter_libelles.py — COLLECTE EXHAUSTIVE des libellés (rapide).

PERFORMANCE : en mode read_only, openpyxl est rapide pour lire les lignes EN
SÉQUENCE (iter_rows) mais très lent pour les accès dispersés (ws.cell). Ce
collecteur lit donc chaque ligne UNE fois, sous forme de tableau de valeurs,
et y trouve libellé + 1re valeur voisine sans aucun accès cellule ponctuel.
"""
import warnings; warnings.filterwarnings("ignore")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

FEUILLES_DEFAUT = ["InpC", "InpS", "InpS-M", "Summary", "Results", "Output", "Oper"]


def _num(v):
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def collecter(fichier, feuilles=FEUILLES_DEFAUT, max_lignes=None, lmin=4, lmax=80):
    """Catalogue : [{libelle, cellule_libelle, adresse_valeur, valeur}].
    Lecture 100% séquentielle : chaque ligne est convertie en liste de valeurs,
    puis on cherche dans cette liste (en mémoire) — pas d'accès cellule dispersé."""
    wv = load_workbook(fichier, data_only=True, read_only=True)
    catalogue = []
    for nom in feuilles:
        if nom not in wv.sheetnames:
            continue
        ws = wv[nom]
        for ri, row in enumerate(ws.iter_rows(max_row=max_lignes), start=1):
            vals = [cell.value for cell in row]
            n = len(vals)
            for i, v in enumerate(vals):
                if isinstance(v, str):
                    s = v.strip()
                    if lmin < len(s) < lmax:
                        # 1re valeur numérique à droite, dans la MÊME ligne déjà lue
                        adr_val = val = None
                        for j in range(i + 1, n):
                            if _num(vals[j]):
                                adr_val = f"{nom}!{get_column_letter(j+1)}{ri}"
                                val = vals[j]
                                break
                        if adr_val is not None:
                            catalogue.append({
                                "libelle": s,
                                "cellule_libelle": f"{nom}!{get_column_letter(i+1)}{ri}",
                                "adresse_valeur": adr_val,
                                "valeur": val,
                            })
    return catalogue


if __name__ == "__main__":
    import sys, time
    f = sys.argv[1] if len(sys.argv) > 1 else "kikot.xlsm"
    t = time.time()
    cat = collecter(f)
    print(f"{len(cat)} libellés en {time.time()-t:.1f}s")


    for e in cat[:5]:
        print(f"  {e['cellule_libelle']} «{e['libelle'][:40]}» -> {e['adresse_valeur']}={e['valeur']}")
